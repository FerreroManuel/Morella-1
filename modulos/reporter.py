                                    ############ INICIO DE IMPORTACIONES ############

import funciones_caja as caja
import funciones_rendiciones as rend
import funciones_cuentas as ctas
import funciones_mantenimiento as mant
import funciones_ventas as vent
import correo as email
from fpdf import FPDF
from datetime import datetime, date
import os
import re
import psycopg2 as sql
import psycopg2.errors
from openpyxl import Workbook
from smtplib import SMTPAuthenticationError
from socket import gaierror
from pprint import pprint

                                    ############## FIN DE IMPORTACIONES #############



#################################################################################################################
############################################# REPORT CAJA DIARIA ################################################
#################################################################################################################

def report_caja_diaria(s_final: float | int):
    """Genera un reporte de la caja diaria, lo guarda en formato PDF y
    luego lo abre en el programa predeterminado.

    :param s_final: Saldo final de caja
    :type s_final: float or int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    total_ing = 0
    total_egr = 0
    saldo_inicial = float(caja.obtener_saldo_inicial())
    saldo_final = s_final
    categ_ing = caja.obtener_categ_ing()
    categ_egr = caja.obtener_categ_egr()
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############
    
    
    ############ INICIO DE FUNCIONES ############

    def buscar_imp_reg(categ: str) -> list:
        """Recupera de la base de datos todos los movimientos de caja de una
        categoría específica que no se encuentren en una caja cerrada y los
        retorna en una lista.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: list
        """
        instruccion = f"SELECT * FROM caja WHERE categoria='{categ}' AND cerrada = '0'"
        datos = mant.run_query(instruccion, fetch="all")
        return datos


    def tot_cat_ing(categ: str) -> float | int:
        """Recupera desde la base de datos el total de ingresos de 
        la caja actual, pertenecientes a una categoría de caja 
        específica, y lo retorna.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: float or int
        """
        instruccion = f"SELECT SUM(ingreso) FROM caja WHERE categoria = '{categ}' AND cerrada = '0'"
        ingresos = mant.run_query(instruccion, fetch="one")
        
        if ingresos[0]:
            return ingresos[0]
        
        else:
            return 0


    def tot_cat_egr(categ: str) -> float | int:
        """Recupera desde la base de datos el total de egresos
        pertenecientes a una categoría de caja y lo retorna.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: float or int
        """
        instruccion = f"SELECT SUM(egreso) FROM caja WHERE categoria = '{categ}' AND cerrada = '0'"
        egresos = mant.run_query(instruccion, fetch="one")
        
        if egresos[0]:
            return egresos[0]
        
        else:
            return 0


    def select_categ(categ: list) -> list:
        """Recibe una lista con nombres de categoría de caja y retorna
        una lista que contiene el nombre sólo de aquellas que tengan
        movimientos dentro de la caja actual.

        :param categ: Lista de categorías de caja.
        :type categ: list

        :rtype: list
        """
        lista_categ = []
        
        for i in categ:
            cat = buscar_imp_reg(i)
        
            if cat:
                lista_categ.append(i)
        
        return lista_categ


    def imprimir_registros_ing(categ: str):
        """Escribe en el PDF todos los movimientos de ingreso de
        una caja sin cerrar de una categoría específica.

        Contenido:
        - Nombre de la categoría (Arial Negrita 10p).
        - Una línea por cada movimiento (Arial 10p) conteniendo:
          - Descripción.
          - Número de transacción (ticket, recibo, rendición, etc).
          - Monto.
          - Observaciones.
        - Total de ingresos de la categoría (Arial Negrita 10p).

        :param categ: Nombre de categoría de caja.
        :type categ: str
        """
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'{categ}', 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
    
        for i in buscar_imp_reg(categ):
            i_d, cat, des, tra, ing, egr, obs, dia, mes, año, cer, use = i
            pdf.set_font('Arial', '', 10)
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(50, 5, f'{des}', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, f'{tra}', 0, 0, 'L')
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(10, 5, f'{ing:.2f}', 0, 0, 'R')
            pdf.cell(25, 5, f'', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(0, 5, f'{obs}', 0, 1, 'L')
    
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'$ {tot_cat_ing(categ):.2f}', 0, 0, 'R')
        pdf.cell(-27, 5, f'Total {categ}: ', 0, 1, 'R')
        pdf.ln(2)
        

    def imprimir_registros_egr(categ: str):
        """Escribe en el PDF todos los movimientos de egreso de
        una caja sin cerrar de una categoría específica.

        Contenido:
        - Nombre de la categoría (Arial Negrita 10p).
        - Una línea por cada movimiento (Arial 10p) conteniendo:
          - Descripción.
          - Número de transacción (ticket, recibo, rendición, etc).
          - Monto.
          - Observaciones.
        - Total de egresos de la categoría (Arial Negrita 10p).

        :param categ: Nombre de categoría de caja.
        :type categ: str
        """
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'{categ}', 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
        for i in buscar_imp_reg(categ):
            i_d, cat, des, tra, ing, egr, obs, dia, mes, año, cer, use = i
            pdf.set_font('Arial', '', 10)
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(50, 5, f'{des}', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, f'{tra}', 0, 0, 'L')
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, f'', 0, 0, 'L')
            pdf.cell(20, 5, f'({egr:.2f})', 0, 0, 'R')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(0, 5, f'{obs}', 0, 1, 'L')
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'($ {tot_cat_egr(categ):.2f})', 0, 0, 'R')
        pdf.cell(-27, 5, f'Total {categ}: ', 0, 1, 'R')
        pdf.ln(2)


    def total_ingresos(saldo_inicial: float | int) -> float | int:
        """Recibe el saldo inicial de la caja, lo suma al total de ingresos
        de la caja actual y retorna el resultado.

        :param saldo_inicial: Saldo inicial de la caja.
        :type saldo_inicial: float or int

        :rtype: float or int
        """
        instruccion = "SELECT SUM(ingreso) FROM caja WHERE cerrada = 0"
        tot_ingresos = mant.run_query(instruccion, fetch="one")[0]
        
        if tot_ingresos == None:
            tot_ingresos = 0

        return tot_ingresos + saldo_inicial


    def total_egresos(saldo_final: float | int) -> float | int:
        """Recibe el saldo final de la caja, lo suma al total de egresos
        de la caja actual y retorna el resultado.

        :param saldo_inicial: Saldo final de la caja.
        :type saldo_inicial: float or int

        :rtype: float or int
        """
        instruccion = "SELECT SUM(egreso) FROM caja WHERE cerrada = 0"
        tot_egresos = mant.run_query(instruccion, fetch="one")[0]

        if tot_egresos == None:
            tot_egresos = 0

        return tot_egresos + saldo_final

    ############ FIN DE FUNCIONES ############

    ############ INICIO DE VARIABLES DEPENDIENTES ############

    total_ing = total_ingresos(saldo_inicial)
    total_egr = total_egresos(saldo_final)
    final = total_ing - total_egr
    contador = caja.obtener_contador()

    ########### FIN DE VARIABLES DEPENDIENTES ############
    
    ############ INICIO DE REPORT ############

    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha, hora y número de caja (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Descripción
              - N°Transacción
              - Ingreso
              - Egreso
              - Observaciones
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'CAJA DIARIA', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # N° de cierre
            self.cell(-74, 35, f'Número de cierre: {str(contador).rjust(6, "0")}', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.set_font('Arial', 'B', 10)
            self.cell(45, 5, 'Descripción', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(20, 5, 'N°Transacción', 0, 0, 'L')
            pdf.cell(10, 5, '', 0, 0, 'L')
            self.cell(15, 5, 'Ingreso', 0, 0, 'L')
            pdf.cell(20, 5, 'Egreso', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(0, 5, 'Observaciones', 0, 1, 'L')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()
    
    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 5, 'CAJA INICIAL: ', 0, 0, "L")
    pdf.cell(0, 5, f'$ {saldo_inicial:.2f}', 0, 1, 'R')
    pdf.set_font('Arial', '', 10)
    
    for i in select_categ(categ_ing):
        imprimir_registros_ing(i)
    
    for e in select_categ(categ_egr):
        imprimir_registros_egr(e)
    
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 5, 'CIERRE DE CAJA: ', 0, 0, "L")
    pdf.cell(0, 5, f'($ {saldo_final:.2f})', 0, 1, 'R')
    pdf.ln(15)
    pdf.cell(0, 5, f' {total_ing:.2f} ', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL INGRESOS: $ ', 0, 1, 'R')
    pdf.cell(0, 5, f'( {total_egr:.2f})', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL EGRESOS: $ ', 0, 1, 'R')
    pdf.cell(0, 5, f'_____________________________', 0, 1, 'R')
    
    if final == 0:
        pdf.cell(0, 5, f' {final:.2f} ', 0, 0, 'R')
        pdf.cell(-27, 5, f'DIFERENCIA: $ ', 0, 1, 'R')
    
    elif final > 0:
        pdf.cell(0, 5, f' {final:.2f} ', 0, 0, 'R')
        pdf.cell(-27, 5, f'FALTANTE: $ ', 0, 1, 'R')
    
    elif final < 0:
        pdf.cell(0, 5, f' {-1*final:.2f}', 0, 0, 'R')
        pdf.cell(-27, 5, f'SOBRANTE: $ ', 0, 1, 'R')
    
    pdf.ln(2)

    pdf.output(mant.re_path(f'reports/caja/diaria/caja_diaria-{str(contador).rjust(6, "0")}.pdf'), 'F')


    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Luego de cerrarlo presione enter para continuar...")
    ruta = mant.re_path('reports/caja/diaria')
    arch = f'caja_diaria-{str(contador).rjust(6, "0")}.pdf'

    os.chdir(ruta)
    os.system(arch)

    ruta = mant.MODULES_DIR
    os.chdir(ruta)

############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################## REPORT CAJA MENSUAL DETALLADA ##########################################
#################################################################################################################

def report_caja_mensual_det(mes: int, año: int):
    """Genera un reporte de la caja mensual, detallando cada movimiento de cada
    categoría, luego lo guarda en formato PDF y lo abre en el programa predeterminado.

    :param mes: Mes a imprimir
    :type mes: int

    :param año: Año a imprimir (hasta 2099 se aceptan dos dígitos, hasta 2999 se aceptan tres dígitos)
    :type año: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    mes = f'{mes}'.rjust(2, '0')
    año = f'{año}'.rjust(3, '0').rjust(4, '2')
    total_ing = 0
    total_egr = 0
    categ_ing = caja.obtener_categ_ing()
    categ_egr = caja.obtener_categ_egr()
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############
    
    
    ############ INICIO DE FUNCIONES ############

    def buscar_imp_reg(categ: str) -> list:
        """Recupera de la base de datos todos los movimientos de caja de una
        categoría específica en un mes específico y los retorna en una lista.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: list
        """
        instruccion = f"SELECT * FROM caja WHERE categoria='{categ}' AND mes='{mes}' AND año='{año}'"
        datos = mant.run_query(instruccion, fetch="all")

        return datos


    def tot_cat_ing(categ: str) -> float | int:
        """Recupera desde la base de datos el total de ingresos pertenecientes
        a una categoría de caja específica, en un mes específico, y lo retorna.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: float or int
        """
        instruccion = f"SELECT SUM(ingreso) FROM caja WHERE categoria='{categ}' AND mes='{mes}' AND año='{año}'"
        ingresos = mant.run_query(instruccion, fetch="one")
        
        if ingresos[0]:
            return ingresos[0]
        
        else:
            return 0


    def tot_cat_egr(categ: str) -> float | int:
        """Recupera desde la base de datos el total de egresos pertenecientes
        a una categoría de caja específica, en un mes específico, y lo retorna.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: float or int
        """
        instruccion = f"SELECT SUM(egreso) FROM caja WHERE categoria='{categ}' AND mes='{mes}' AND año='{año}'"
        egresos = mant.run_query(instruccion, fetch="one")
        
        if egresos[0]:
            return egresos[0]
        
        else:
            return 0


    def select_categ(categ: list) -> list:
        """Recibe una lista con nombres de categoría de caja y retorna
        una lista que contiene el nombre sólo de aquellas que tengan
        movimientos dentro del mes indicado.

        :param categ: Lista de categorías de caja.
        :type categ: list

        :rtype: list
        """
        lista_categ = []
        
        for i in categ:
            cat = buscar_imp_reg(i)
        
            if cat:
                lista_categ.append(i)
        
        return lista_categ


    def imprimir_registros_ing_mes(categ: str):
        """Escribe en el PDF todos los movimientos de ingreso de
        caja, en un mes específico, de una categoría específica.

        Contenido:
        - Nombre de la categoría (Arial Negrita 10p).
        - Una línea por cada movimiento (Arial 10p) conteniendo:
          - Descripción.
          - Número de transacción (ticket, recibo, rendición, etc).
          - Monto.
          - Observaciones.
          - Fecha (MM/AAAA).
        - Total de ingresos de la categoría (Arial Negrita 10p).

        :param categ: Nombre de categoría de caja.
        :type categ: str
        """
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'{categ}', 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
        
        for i in buscar_imp_reg(categ):
            i_d, cat, des, tra, ing, egr, obs, dia, mes, año, cer, use = i
            pdf.set_font('Arial', '', 10)
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(50, 5, f'{des}', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, f'{tra}', 0, 0, 'L')
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(10, 5, f'{ing:.2f}', 0, 0, 'R')
            pdf.cell(25, 5, f'', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(60, 5, f'{obs}', 0, 0, 'L')
            pdf.cell(0, 5 , f'{mes}/{año}', 0, 1, 'L')
        
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'$ {tot_cat_ing(categ):.2f}', 0, 0, 'R')
        pdf.cell(-27, 5, f'Total {categ}: ', 0, 1, 'R')
        pdf.ln(2)
        

    def imprimir_registros_egr_mes(categ: str):
        """Escribe en el PDF todos los movimientos de egreso de
        caja, en un mes específico, de una categoría específica.

        Contenido:
        - Nombre de la categoría (Arial Negrita 10p).
        - Una línea por cada movimiento (Arial 10p) conteniendo:
          - Descripción.
          - Número de transacción (ticket, recibo, rendición, etc).
          - Monto.
          - Observaciones.
          - Fecha (MM/AAAA).
        - Total de egresos de la categoría (Arial Negrita 10p).

        :param categ: Nombre de categoría de caja.
        :type categ: str
        """
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'{categ}', 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
        
        for i in buscar_imp_reg(categ):
            i_d, cat, des, tra, ing, egr, obs, dia, mes, año, cer, use = i
            pdf.set_font('Arial', '', 10)
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(50, 5, f'{des}', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, f'{tra}', 0, 0, 'L')
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, f'', 0, 0, 'L')
            pdf.cell(20, 5, f'({egr:.2f})', 0, 0, 'R')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(60, 5, f'{obs}', 0, 0, 'L')
            pdf.cell(0, 5 , f'{mes}/{año}', 0, 1, 'L')
        
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'($ {tot_cat_egr(categ):.2f})', 0, 0, 'R')
        pdf.cell(-27, 5, f'Total {categ}: ', 0, 1, 'R')
        pdf.ln(2)


    def total_ingresos_mensual() -> float | int:
        """Recupera de la base de datos el total de ingresos de un
        mes específico y lo retorna.

        :rtype: float or int
        """
        instruccion = f"SELECT SUM (ingreso) FROM caja WHERE mes = '{mes}' AND año = '{año}'"
        tot_ingresos = mant.run_query(instruccion, fetch="one")[0]
        
        if tot_ingresos:
            return tot_ingresos
        
        else:
            return 0


    def total_egresos_mensual() -> float | int:
        """Recupera de la base de datos el total de ingresos de un
        mes específico y lo retorna.

        :rtype: float or int
        """
        instruccion = f"SELECT SUM (egreso) FROM caja WHERE mes = '{mes}' AND año = '{año}'"
        tot_egresos = mant.run_query(instruccion, fetch="one")[0]
        
        if tot_egresos:
            return tot_egresos
        
        else:
            return 0


    def str_mes(mes: str) -> str:
        """Recibe una cadena conteniendo el mes como número de dos dígitos
        y retorna una cadena con el nombre correspondiente.

        :param mes: Mes (número dos dígitos).
        :type mes: str

        :rtype: str
        """
        if mes == '01':
            string_mes = 'Enero'
        
        if mes == '02':
            string_mes = 'Febrero'
        
        elif mes == '03':
            string_mes = 'Marzo'
        
        elif mes == '04':
            string_mes = 'Abril'
        
        elif mes == '05':
            string_mes = 'Mayo'
        
        elif mes == '06':
            string_mes = 'Junio'
        
        elif mes == '07':
            string_mes = 'Julio'
        
        elif mes == '08':
            string_mes = 'Agosto'
        
        elif mes == '09':
            string_mes = 'Septiembre'
        
        elif mes == '10':
            string_mes = 'Octubre'
        
        elif mes == '11':
            string_mes = 'Noviembre'
        
        elif mes == '12':
            string_mes = 'Diciembre'
        
        return string_mes

    ############ FIN DE FUNCIONES ############

    ############ INICIO DE VARIABLES DEPENDIENTES ############

    string_mes = str_mes(mes)
    total_ing = total_ingresos_mensual()
    total_egr = total_egresos_mensual()
    
    ########### FIN DE VARIABLES DEPENDIENTES ############

    ############ INICIO DE REPORT ############

    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Descripción
              - N°Transacción
              - Ingreso
              - Egreso
              - Observaciones
              - Fecha
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'CAJA MENSUAL DETALLADA', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            self.set_font('Arial', 'B', 10)
            # Mes de cierre
            self.cell(-81, 35, f'{string_mes} de {año}', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.cell(45, 5, 'Descripción', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(20, 5, 'N°Transacción', 0, 0, 'L')
            pdf.cell(10, 5, '', 0, 0, 'L')
            self.cell(15, 5, 'Ingreso', 0, 0, 'L')
            pdf.cell(20, 5, 'Egreso', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(60, 5, 'Observaciones', 0, 0, 'L')
            pdf.cell(0, 5, 'Fecha', 0, 1, 'L')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('Arial', '', 10)
    
    for i in select_categ(categ_ing):
        imprimir_registros_ing_mes(i)
    
    for e in select_categ(categ_egr):
        imprimir_registros_egr_mes(e)
    
    pdf.ln(15)
    pdf.cell(0, 5, f' {total_ing:.2f} ', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL INGRESOS: $ ', 0, 1, 'R')
    pdf.cell(0, 5, f'( {total_egr:.2f})', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL EGRESOS: $ ', 0, 1, 'R')
    pdf.ln(2)
    
    pdf.output(mant.re_path(f'reports/caja/mensual/detallada/caja_{str.lower(string_mes)}-{año}.pdf'), 'F')


    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Cierre el archivo para continuar...")
    
    ruta = mant.re_path('reports/caja/mensual/detallada/')
    arch = f'caja_{str.lower(string_mes)}-{año}.pdf'
    
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)

############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################### REPORT CAJA MENSUAL COMPRIMIDA ########################################
#################################################################################################################

def report_caja_mensual_comp(mes: int, año: int):
    """Genera un reporte de la caja mensual, mostrando sólo los totales de cada
    categoría, luego lo guarda en formato PDF y lo abre en el programa predeterminado.

    :param mes: Mes a imprimir
    :type mes: int

    :param año: Año a imprimir (hasta 2099 se aceptan dos dígitos, hasta 2999 se aceptan tres dígitos)
    :type año: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    mes = f'{mes}'.rjust(2, '0')
    año = f'{año}'.rjust(3, '0').rjust(4, '2')
    total_ing = 0
    total_egr = 0
    categ_ing = caja.obtener_categ_ing()
    categ_egr = caja.obtener_categ_egr()
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############
    
    
    ############ INICIO DE FUNCIONES ############

    def buscar_imp_reg(categ: str) -> list:
        """Recupera de la base de datos todos los movimientos de caja de una
        categoría específica en un mes específico y los retorna en una lista.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: list
        """
        instruccion = f"SELECT * FROM caja WHERE categoria='{categ}' AND mes='{mes}' AND año='{año}'"
        datos = mant.run_query(instruccion, fetch="all")
        return datos


    def tot_cat_ing(categ: str) -> float | int:
        """Recupera desde la base de datos el total de ingresos pertenecientes
        a una categoría de caja específica, en un mes específico, y lo retorna.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: float or int
        """
        instruccion = f"SELECT SUM(ingreso) FROM caja WHERE categoria='{categ}' AND mes='{mes}' AND año='{año}'"
        ingresos = mant.run_query(instruccion, fetch="one")
        
        if ingresos[0]:
            return ingresos[0]
        
        else:
            return 0


    def tot_cat_egr(categ: str) -> float | int:
        """Recupera desde la base de datos el total de egresos pertenecientes
        a una categoría de caja específica, en un mes específico, y lo retorna.

        :param categ: Categoría de caja
        :type categ: str

        :rtype: float or int
        """
        instruccion = f"SELECT SUM(egreso) FROM caja WHERE categoria='{categ}' AND mes='{mes}' AND año='{año}'"
        egresos = mant.run_query(instruccion, fetch="one")
        
        if egresos[0]:
            return egresos[0]
        
        else:
            return 0


    def select_categ(categ: list) -> list:
        """Recibe una lista con nombres de categoría de caja y retorna
        una lista que contiene el nombre sólo de aquellas que tengan
        movimientos dentro de la caja actual.

        :param categ: Lista de categorías de caja.
        :type categ: list

        :rtype: list
        """
        lista_categ = []
        
        for i in categ:
            cat = buscar_imp_reg(i)
        
            if cat:
                lista_categ.append(i)
        
        return lista_categ


    def imprimir_registros_ing_mes(categ: str):
        """Escribe en el PDF el total de ingresos de una categoría específica.

        Contenido:
        - Nombre de la categoría (Arial 10p).
        - Total de ingresos de la categoría (Arial 10p).

        :param categ: Nombre de categoría de caja.
        :type categ: str
        """
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 5, f'{categ}', 0, 0, 'L')
        pdf.cell(-58, 5, f'$ {tot_cat_ing(categ):.2f}', 0, 1, 'R')
        pdf.ln(2)
        

    def imprimir_registros_egr_mes(categ: str):
        """Escribe en el PDF el total de egresos de una categoría específica.

        Contenido:
        - Nombre de la categoría (Arial 10p).
        - Total de egresos de la categoría (Arial 10p).

        :param categ: Nombre de categoría de caja.
        :type categ: str
        """
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 5, f'{categ}', 0, 0, 'L')
        pdf.cell(0, 5, f'($ {tot_cat_egr(categ):.2f})', 0, 1, 'R')
        pdf.ln(2)


    def total_ingresos_mensual() -> float | int:
        """Recupera de la base de datos el total de ingresos de un
        mes específico y lo retorna.

        :rtype: float or int
        """
        instruccion = f"SELECT SUM (ingreso) FROM caja WHERE mes = '{mes}' AND año = '{año}'"
        tot_ingresos = mant.run_query(instruccion, fetch="one")[0]
        
        if tot_ingresos:
            return tot_ingresos
        
        else:
            return 0


    def total_egresos_mensual() -> float | int:
        """Recupera de la base de datos el total de ingresos de un
        mes específico y lo retorna.

        :rtype: float or int
        """
        instruccion = f"SELECT SUM (egreso) FROM caja WHERE mes = '{mes}' AND año = '{año}'"
        tot_egresos = mant.run_query(instruccion, fetch="one")[0]
        
        if tot_egresos:
            return tot_egresos
        
        else:
            return 0


    def str_mes(mes: str) -> str:
        """Recibe una cadena conteniendo el mes como número de dos dígitos
        y retorna una cadena con el nombre correspondiente.

        :param mes: Mes (número dos dígitos).
        :type mes: str

        :rtype: str
        """
        if mes == '01':
            string_mes = 'Enero'
        
        if mes == '02':
            string_mes = 'Febrero'
        
        elif mes == '03':
            string_mes = 'Marzo'
        
        elif mes == '04':
            string_mes = 'Abril'
        
        elif mes == '05':
            string_mes = 'Mayo'
        
        elif mes == '06':
            string_mes = 'Junio'
        
        elif mes == '07':
            string_mes = 'Julio'
        
        elif mes == '08':
            string_mes = 'Agosto'
        
        elif mes == '09':
            string_mes = 'Septiembre'
        
        elif mes == '10':
            string_mes = 'Octubre'
        
        elif mes == '11':
            string_mes = 'Noviembre'
        
        elif mes == '12':
            string_mes = 'Diciembre'
        
        return string_mes

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    string_mes = str_mes(mes)
    total_ing = total_ingresos_mensual()
    total_egr = total_egresos_mensual()
    
    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############

    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Categoría
              - Ingreso
              - Egreso
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'CAJA MENSUAL COMPRIMIDA', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            self.set_font('Arial', 'B', 10)
            # Mes de cierre
            self.cell(-81, 35, f'{string_mes} de {año}', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.cell(45, 5, 'Categoría', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(20, 5, '', 0, 0, 'L')
            pdf.cell(10, 5, '', 0, 0, 'L')
            self.cell(15, 5, '', 0, 0, 'L')
            pdf.cell(20, 5, '', 0, 0, 'L')
            pdf.cell(7, 5, '', 0, 0, 'L')
            pdf.cell(55, 5, 'Ingreso', 0 , 0, 'L')
            pdf.cell(0, 5, 'Egreso', 0, 1, 'L')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            #self.set_font('Arial', 'BI', 8)
            #self.cell(0, 10, 'Solutions.', 0, 0, 'R')
            #self.set_font('Arial', 'BIU', 8)
            #self.cell(-15, 10, 'MF!', 0, 0, 'R')
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()
    
    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('Arial', '', 10)
    
    for i in select_categ(categ_ing):
        imprimir_registros_ing_mes(i)
    
    for e in select_categ(categ_egr):
        imprimir_registros_egr_mes(e)
    
    pdf.ln(15)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 5, f' {total_ing:.2f} ', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL INGRESOS: $ ', 0, 1, 'R')
    pdf.cell(0, 5, f'( {total_egr:.2f})', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL EGRESOS: $ ', 0, 1, 'R')
    pdf.ln(2)
    
    pdf.output(mant.re_path(f'reports/caja/mensual/comprimida/caja_{str.lower(string_mes)}-{año}-COMP.pdf'), 'F')


    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Cierre el archivo para continuar...")
    
    ruta = mant.re_path('reports/caja/mensual/comprimida/')
    arch = f'caja_{str.lower(string_mes)}-{año}-COMP.pdf'
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)

############################################### FIN DE REPORT ###################################################




#################################################################################################################
####################################### REPORT CAJA MENSUAL POR COBRADOR ########################################
#################################################################################################################

def report_caja_mensual_por_cob(mes: int, año: int):
    """Genera un reporte de la caja mensual, detallando todos los ingresos de cada
    cobrador, luego lo guarda en formato PDF y lo abre en el programa predeterminado.

    :param mes: Mes a imprimir
    :type mes: int

    :param año: Año a imprimir (hasta 2099 se aceptan dos dígitos, hasta 2999 se aceptan tres dígitos)
    :type año: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    mes = f'{mes}'.rjust(2, '0')
    año = f'{año}'.rjust(3, '0').rjust(4, '2')
    cobradores = caja.obtener_cobrador()
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############
    
    
    ############ INICIO DE FUNCIONES ############

    def buscar_imp_reg_por_cob(cobrador: str, mes: str, año: str) -> list:
        """Recupera de la base de datos todos los ingresos de caja de un cobrador
        específico en un mes específico y los retorna en una lista.

        :param cobrador: Nombre de cobrador
        :type categ: str

        :param mes: Mes (cadena, dos dígitos)
        :type mes: str

        :param año: Año (cadena, cuatro dígitos)
        :type año: str

        :rtype: list
        """
        instruccion = f"SELECT * FROM caja WHERE descripcion='{cobrador}' AND mes='{mes}' AND año='{año}'"
        datos = mant.run_query(instruccion, fetch="all")

        return datos


    def select_cob(cobradores: list, mes: str, año: str) -> list:
        """Recibe una lista con los nombres delos cobradores y retorna
        una lista que contiene el nombre sólo de aquellos que tengan
        ingresos registrados dentro del mes indicado.

        :param cobradores: Lista de nombres de cobradores.
        :type cobradores: list

        :param mes: Mes (cadena, dos dígitos).
        :type mes: str

        :param año: Año (cadena, cuatro dígitos).
        :type año: str

        :rtype: list
        """
        lista_cob = []

        for i in cobradores:
            cob = buscar_imp_reg_por_cob(i, mes, año)
            
            if cob:
                lista_cob.append(i)

        return lista_cob


    def imprimir_registros_por_cob(cobrador: str, mes: str, año: str):
        """Escribe en el PDF todos los movimientos de ingreso de un cobrador
        en un mes específico.

        Contenido:
        - Nombre del cobrador (Arial Negrita 10p).
        - Una línea por cada movimiento (Arial 10p) conteniendo:
          - Panteón (categoría).
          - Número de transacción (rendición).
          - Monto.
          - Observaciones.
          - Fecha (MM/AAAA).
        - Total de ingresos del cobrador (Arial Negrita 10p).

        :param cobrador: Nombre del cobrador.
        :type cobrador: str
        """
        tot_i_por_cob = caja.total_ing_por_cob(cobrador, mes, año)


        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'{cobrador}', 0, 1, 'L')
        pdf.set_font('Arial', '', 10)
        
        for i in buscar_imp_reg_por_cob(cobrador, mes, año):
            i_d, cat, des, tra, ing, egr, obs, dia, mes, año, cer, use = i
        
            if egr:
                continue

            if cat not in caja.obtener_panteon():
                continue

            pdf.set_font('Arial', '', 10)
            pdf.cell(5, 5, '', 0, 0, 'L')
            pdf.cell(60, 5, f'{cat}', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(10, 5, f'{tra}', 0, 0, 'L')
            pdf.cell(15, 5, '', 0, 0, 'L')
            pdf.cell(10, 5, f'{ing}', 0, 0, 'R')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(73, 5, f'{obs}', 0, 0, 'L')
            pdf.cell(0, 5 , f'{mes}/{año}', 0, 1, 'L')
        
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 5, f'$ {tot_i_por_cob:.2f}', 0, 0, 'R')
        pdf.cell(-27, 5, f'Total {cobrador}: ', 0, 1, 'R')
        pdf.ln(2)


    def total_mensual_cobradores(mes: str, año: str) -> float | int:
        """Suma todos los ingresos de cobradores en un mes específico y lo retorna.

        :param mes: Mes (cadena, dos dígitos)
        :type mes: str

        :param año: Año (cadena, cuatro dígitos)
        :type año: str

        :rtype: float or int
        """
        total_cobradores = 0

        for i in cobradores:
            ing_por_cob = caja.total_ing_por_cob(i, mes, año)
            total_cobradores += ing_por_cob

        return total_cobradores


    def str_mes(mes: str) -> str:
        """Recibe una cadena conteniendo el mes como número de dos dígitos
        y retorna una cadena con el nombre correspondiente.

        :param mes: Mes (número dos dígitos).
        :type mes: str

        :rtype: str
        """
        if mes == '01':
            string_mes = 'Enero'
        
        if mes == '02':
            string_mes = 'Febrero'
        
        elif mes == '03':
            string_mes = 'Marzo'
        
        elif mes == '04':
            string_mes = 'Abril'
        
        elif mes == '05':
            string_mes = 'Mayo'
        
        elif mes == '06':
            string_mes = 'Junio'
        
        elif mes == '07':
            string_mes = 'Julio'
        
        elif mes == '08':
            string_mes = 'Agosto'
        
        elif mes == '09':
            string_mes = 'Septiembre'
        
        elif mes == '10':
            string_mes = 'Octubre'
        
        elif mes == '11':
            string_mes = 'Noviembre'
        
        elif mes == '12':
            string_mes = 'Diciembre'
        
        return string_mes


    ############ FIN DE FUNCIONES ############

    ############ INICIO DE VARIABLES DEPENDIENTES ############

    string_mes = str_mes(mes)
    total_mes_cob = total_mensual_cobradores(mes, año)

    ########### FIN DE VARIABLES DEPENDIENTES ############
    
    ############ INICIO DE REPORT ############

    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Panteón (categoría).
              - N°Transacción (rendición).
              - Ingreso.
              - Observaciones.
              - Fecha.
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'CAJA MENSUAL POR COBRADOR', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            self.set_font('Arial', 'B', 10)
            # Mes de cierre
            self.cell(-81, 35, f'{string_mes} de {año}', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.cell(53, 5, 'Panteón', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(20, 5, 'N°Transacción', 0, 0, 'L')
            pdf.cell(12, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, 'Ingreso', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(75, 5, 'Observaciones', 0, 0, 'L')
            pdf.cell(0, 5, 'Fecha', 0, 1, 'L')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()
    
    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font('Arial', '', 10)
    datos = select_cob(cobradores, mes, año)
    
    for i in datos:
        imprimir_registros_por_cob(i, mes, año)
        counter += 1
        mant.barra_progreso(counter, len(datos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    pdf.ln(15)
    pdf.cell(0, 5, f' {total_mes_cob:.2f} ', 0, 0, 'R')
    pdf.cell(-27, 5, f'TOTAL: $ ', 0, 1, 'R')
    pdf.ln(2)
    
    pdf.output(mant.re_path(f'reports/caja/mensual/por_cobrador/caja_{str.lower(string_mes)}-{año}-por_cobrador.pdf'), 'F')
        

    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print()
    print()
    print("Abriendo reporte. Cierre el archivo para continuar...")
    
    ruta = mant.re_path('reports/caja/mensual/por_cobrador/')
    arch = f'caja_{str.lower(string_mes)}-{año}-por_cobrador.pdf'
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)

############################################### FIN DE REPORT ###################################################




#################################################################################################################
################################################### RECIBOS #####################################################
#################################################################################################################

def recibos(cobrador: int, facturacion: str, periodo: str, año: str, reimpresion: bool = False):
    """Genera un reporte en PDF que contiene todos los recibos correspondientes a un cobrador,
    una facturación y un período específicos, luego lo guarda y lo abre con el programa 
    predeterminado.
    
    Los recibos se generan con datos y valores actualizados.

    En caso que reciba el parámetro `reimpresion` como verdadero, se guarda como tal.

    :param cobrador: ID del cobrador.
    :type cobrador: int

    :param facturacion: Tipo de facturación (bicon o nob).
    :type facturacion: str

    :param periodo: Período a imprimir.
    :type periodo: str

    :param año: Año del período a imprimir.
    :type año: str

    :param reimpresion: Indica si los recibos se imprimen durante emisión o reimpresión.
    :type reimpresion: bool
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    counter = 0
    nco = ""
    per = ""

    instruccion = f"""\
        SELECT
            rec.nro_recibo,
            rec.operacion,
            rec.periodo,
            rec.año,
            ops.cobrador,
            cob.cobrador,
            ops.ruta,
            ops.cuotas_favor,
            ops.nombre_alt,
            ops.domicilio_alt,
            soc.nro_socio,
            soc.nombre,
            soc.domicilio,
            soc.localidad,
            soc.cod_postal,
            nic.codigo,
            pan.panteon,
            nic.piso,
            nic.fila,
            nic.numero,
            cat.categoria,
            cat.valor_mant_bicon,
            cat.valor_mant_nob
        FROM
            recibos rec
            JOIN operaciones ops ON rec.operacion = ops.id
            JOIN cobradores cob ON ops.cobrador = cob.id
            JOIN socios soc ON ops.socio = soc.nro_socio
            JOIN nichos nic ON ops.nicho = nic.codigo
            JOIN panteones pan ON nic.panteon = pan.id
            JOIN cat_nichos cat ON nic.categoria = cat.id
        WHERE
            ops.paga = 1
            AND soc.activo = 1
            AND rec.pago = 0
            AND ops.cobrador = {cobrador}
            AND ops.nicho IS NOT NULL
            AND rec.periodo = '{periodo}'
            AND rec.año = '{año}'
        ORDER BY
            ops.ruta,
            ops.socio,
            ops.id,
            rec.nro_recibo
        """

    try:
        recibos = mant.run_query(instruccion, fetch="all")

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############


    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############

    # Logo NOB
    class PDF(FPDF):
        """Clase dedicada únicamente a los recibos de operaciones
        que pertenecen a NOB.
        """
        def header(self):
            """Inserta un escudo de NOB en la esquina superior izquierda
            de cada uno de los recibos de la hoja.

            Inserta ocho escudos sin importar si en la hoja quedan
            recibos en blanco.
            """
            self.image(mant.re_path('docs/logo_nob.jpg'), 11, 4, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 108, 4, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 11, 78, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 108, 78, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 11, 152, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 108, 152, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 11, 226, 10)
            self.image(mant.re_path('docs/logo_nob.jpg'), 108, 226, 10)

    pdf = PDF() if facturacion == 'nob' else FPDF()
    
    pdf.set_margins(10, 0, 10)
    pdf.set_auto_page_break(True, 0)
    pdf.alias_nb_pages()
    pdf.add_page()

    for recibo in recibos:
        ( ndr, id_o, per, año, cob, nco, rut, c_f, nom_alt, dom_alt, nro, nom,
        dom, loc, c_p, cod, pant, pis, fil, num, cat, val_mant_bic, val_mant_nob ) = recibo

        nombre = nom_alt or nom if len(nom_alt or nom) < 42 else (nom_alt or nom)[:39] + '...'
        domicilio = dom_alt or dom if len(dom_alt or dom) < 53 else (dom_alt or dom)[:50] + '...'

        val_mant = val_mant_bic if facturacion == 'bicon' else val_mant_nob

        # Header
        if facturacion == 'bicon':
            # Línea 1
            pdf.set_font('Arial', 'I', 7)
            pdf.cell(190, 3.1, '', 0, 1, 'L')
            pdf.cell(71, 3, '', 0, 0, 'L')
            pdf.cell(25, 3, 'Talón para control', 0, 0, 'L')
            pdf.cell(61, 3, '', 0, 0, 'L')
            pdf.cell(31, 3, 'Talón para el contribuyente', 0, 1, 'L')
            pdf.ln(1)
            
            # Línea 2
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(7, 3, '', 0, 0, 'L')
            pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 0, 'L')
            pdf.cell(7, 3, '', 0, 0, 'L')
            pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 1, 'L')
            
            # Línea 3
            pdf.cell(22, 3, '', 0, 0, 'L')
            pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 0, 'L')
            pdf.cell(18, 3, '', 0, 0, 'L')
            pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 1, 'L')
            
            # Línea 4
            pdf.cell(19, 3, '', 0, 0, 'L')
            pdf.cell(45, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
            pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
            pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
            pdf.cell(13, 3, '', 0, 0, 'L')
            pdf.cell(55, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
            pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
            pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
            pdf.ln(1)

        elif facturacion == 'nob':
            # Línea 1
            pdf.set_font('Arial', 'I', 7)
            pdf.cell(190, 3.1, '', 0, 1, 'L')
            pdf.ln(1)
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(60, 3, "Club Atlético Newell's Old Boys", 0, 0, 'L')
            pdf.set_font('Arial', 'I', 7)
            pdf.cell(14, 3, 'Talón para control', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(24, 3, '', 0, 0, 'L')
            pdf.cell(49, 3, "Club Atlético Newell's Old Boys", 0, 0, 'L')
            pdf.set_font('Arial', 'I', 7)
            pdf.cell(14, 3, 'Talón para el contribuyente', 0, 1, 'L')
            
            # Línea 2
            pdf.set_font('Arial', 'B', 6)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
            pdf.set_font('Arial', 'I', 6)
            pdf.cell(48, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
            pdf.cell(15, 3, '', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 6)
            pdf.cell(19, 3, '', 0, 0, 'L')
            pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
            pdf.set_font('Arial', 'I', 6)
            pdf.cell(47, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
            pdf.cell(15, 3, '', 0, 1, 'L')
            
            # Línea 3
            pdf.set_font('Arial', 'B', 6)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(60, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 0, 'L')
            pdf.cell(38, 3, '', 0, 0, 'L')
            pdf.cell(49, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 1, 'L')
            
            # Línea 4
            pdf.set_font('Arial', '', 6)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(52, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
            pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
            pdf.set_font('Arial', '', 6)
            pdf.cell(17, 3, '', 0, 0, 'L')
            pdf.cell(51, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
            pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
            pdf.ln(1)

        # Línea 1
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(13, 4, 'Socio/a: ', 'LTB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'TB', 0, 'L')
        pdf.cell(68, 4, f'{nombre}', 'TRB', 0, 'L')
        pdf.cell(4, 4, '', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(13, 4, 'Socio/a: ', 'LT', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'T', 0, 'L')
        pdf.cell(68, 4, f'{nombre}', 'TR', 1, 'L')

        # Línea 2
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(95, 1, '', 0, 0, 'L')
        pdf.cell(2, 1, '', 0, 0, 'L')
        pdf.cell(93, 1, '', 'LR', 1, 'L')
        pdf.cell(19, 5, 'Cobrador/a: ', 'LT', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(5, 5, f'{cob}'.rjust(2, '0'), 'T', 0, 'L')
        pdf.cell(39, 5, f'{nco}', 'T', 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(9, 5, 'Ruta:', 'T', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(21, 5, f'{rut}'.rjust(3, '0'), 'RT', 0, 'L')
        pdf.cell(4, 5, '', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(17, 5, 'Domicilio: ', 'L', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(76, 5, f'{domicilio}', 'R', 1, 'L')

        # Línea 3
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(17, 4, 'Categoría: ', 'LB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(76, 4, f'{cat}', 'RB', 0, 'L')
        pdf.cell(4, 4, '', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(17, 4, 'Localidad:', 'LB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(76, 4, f'{loc} - {c_p}', 'BR', 1, 'L')
        pdf.ln(1)

        # Línea 4
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(19, 5, 'Cod. Nicho:', 'LTB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(44, 5, f'{cod}'.rjust(10, '0'), 'TB', 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(7, 5, f'Op:', 'TB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(23, 5, f'{id_o}'.rjust(7, "0"), 'RTB', 0, 'L')
        pdf.cell(4, 5, '', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(16, 5, 'Cobrador:', 'LTB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(5, 5, f'{cob}'.rjust(2, '0'), 'TB', 0, 'L')
        pdf.cell(72, 5, f'{nco}', 'RTB', 1, 'L')
        pdf.ln(1)

        # Línea 5
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(44, 4, 'Panteón', 'LTR', 0, 'C')
        pdf.cell(16, 4, 'Piso', 'LTR', 0, 'C')
        pdf.cell(16, 4, 'Fila', 'LTR', 0, 'C')
        pdf.cell(17, 4, 'Nicho', 'LTR', 0, 'C')
        pdf.cell(4, 4, '', 0, 0, 'C')
        pdf.cell(29, 4, 'Categoría', 'LTR', 0, 'C')
        pdf.cell(34, 4, 'Panteón', 'LTR', 0, 'C')
        pdf.cell(9, 4, 'Piso', 'LTR', 0, 'C')
        pdf.cell(9, 4, 'Fila', 'LTR', 0, 'C')
        pdf.cell(12, 4, 'Nicho', 'LTR', 1, 'C')

        # Línea 6
        pdf.set_font('Arial', '', 9)
        pdf.cell(44, 5, f'{pant}', 'LBR', 0, 'C')
        pdf.cell(16, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
        pdf.cell(16, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
        pdf.cell(17, 5, f'{num}'.rjust(3, '0'), 'LBR', 0, 'C')
        pdf.cell(4, 5, '', 0, 0, 'C')
        pdf.cell(29, 5, f'{cat}', 'LBR', 0, 'C')
        pdf.cell(34, 5, f'{pant}', 'LBR', 0, 'C')
        pdf.cell(9, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
        pdf.cell(9, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
        pdf.cell(12, 5, f'{num}'.rjust(3, '0'), 'LBR', 1, 'C')
        pdf.ln(3)

        # Línea 7
        if per == "Enero - Febrero":
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(44, 5, f'{per} - {int(año)+1}', 'RTB', 0, 'L')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(21, 5, f'$ {val_mant:.2f}', 'RTB', 0, 'R')
            pdf.cell(4, 5, '', 0, 0, 'C')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(44, 5, f'{per} - {int(año)+1}', 'RTB', 0, 'L')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(21, 5, f'$ {val_mant:.2f}', 'RTB', 1, 'R')
            pdf.ln(2)

        elif per == "Diciembre - Enero":
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(44, 5, f'{per} - {año}/{int(año[:2])+1}', 'RTB', 0, 'L')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(21, 5, f'$ {val_mant:.2f}', 'RTB', 0, 'R')
            pdf.cell(4, 5, '', 0, 0, 'C')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(44, 5, f'{per} - {año}/{int(año[:2])+1}', 'RTB', 0, 'L')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(21, 5, f'$ {val_mant:.2f}', 'RTB', 1, 'R')
            pdf.ln(2)

        else:
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(44, 5, f'{per} - {año}', 'RTB', 0, 'L')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(21, 5, f'$ {val_mant:.2f}', 'RTB', 0, 'R')
            pdf.cell(4, 5, '', 0, 0, 'C')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(44, 5, f'{per} - {año}', 'RTB', 0, 'L')
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(21, 5, f'$ {val_mant:.2f}', 'RTB', 1, 'R')
            pdf.ln(2)

        # Línea 8
        q_rec_impagos = len(rend.obtener_recibos_impagos_op(id_o))
        debe = 0

        if c_f < 0:
            debe += abs(c_f)
        debe += q_rec_impagos - 1

        if debe:
            pdf.cell(93, 4, f'----------- ATENCIÓN: El asociado adeuda {debe} cuotas. ----------', 1, 1, 'C')
            # Margen
            pdf.cell(190, 13, ' ', 0, 1, 'L')

        else:
            pdf.cell(190, 17, ' ', 0, 1, 'L')

        counter += 1
        mant.barra_progreso(counter, len(recibos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')

    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')

    try:
        # Crea la carpeta si no existe.
        if not os.path.isdir(mant.re_path(f'reports/recibos/{nco}')):
            os.mkdir(mant.re_path(f'reports/recibos/{nco}'))

        # Evita sobreescribir un archivo existente
        output_counter = 0
        prefix = "reimpresion_" if reimpresion else ""
        periodo = f"{per}_{año}".replace(' ', '')

        output_name = f"{prefix}recibos_{periodo}.pdf"

        while os.path.isfile(mant.re_path(f'reports/recibos/{nco}/{output_name}')):
            output_counter += 1
            output_name = f"{prefix}recibos_{periodo}_({output_counter}).pdf"

        pdf.output(mant.re_path(f'reports/recibos/{nco}/{output_name}'), 'F')

        ############ ABRIR REPORT ############

        ruta = mant.re_path(f'reports/recibos/{nco}/')
        arch = output_name.replace('(', '^(')
        os.chdir(ruta)
        os.system(arch)

        ruta = mant.MODULES_DIR
        os.chdir(ruta)

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print("")
        print()
        return

############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################## LISTADO DE RECIBOS ###############################################
#################################################################################################################

def listado_recibos(cobrador: int, facturacion: str, periodo: str, año: str, ops_arregladas: list, reimpresion: bool = False):
    """Genera un reporte en PDF que contiene un listado con la información de todos
    los recibos correspondientes a un cobrador, una facturación y un período específicos,
    luego lo guarda y lo abre con el programa predeterminado.

    El listado se reimprime con datos y valores actualizados.

    En caso que reciba el parámetro `reimpresion` como verdadero, se indica en el título
    y se guarda como tal.

    :param cobrador: ID del cobrador.
    :type cobrador: int

    :param facturacion: Tipo de facturación (bicon o nob).
    :type facturacion: str

    :param periodo: Período a imprimir.
    :type periodo: str

    :param año: Año del período a imprimir.
    :type año: str

    :param ops_arregladas: Operaciones que se les necesitó arreglar sus cuotas a favor.
    :type ops_arregladas: int

    :param reimpresion: Indica si los recibos se imprimen durante emisión o reimpresión.
    :type reimpresion: bool
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    imp_acu = float(0)
    nco = caja.obtener_nom_cobrador(cobrador)
    per = ""
    title = "LISTADO DE RECIBOS EMITIDOS" + (" (REIMPRESIÓN)" if reimpresion else "")

    instruccion = f"""\
        SELECT
            rec.operacion,
            rec.periodo,
            rec.año,
            ops.ruta,
            ops.cuotas_favor,
            ops.nombre_alt,
            ops.domicilio_alt,
            soc.nro_socio,
            soc.nombre,
            soc.domicilio,
            cat.valor_mant_bicon,
            cat.valor_mant_nob
        FROM
            recibos rec
            JOIN operaciones ops ON rec.operacion = ops.id
            JOIN socios soc ON ops.socio = soc.nro_socio
            JOIN nichos nic ON ops.nicho = nic.codigo
            JOIN cat_nichos cat ON nic.categoria = cat.id
        WHERE
            ops.paga = 1
            AND soc.activo = 1
            AND rec.pago = 0
            AND ops.cobrador = {cobrador}
            AND ops.nicho IS NOT NULL
            AND rec.periodo = '{periodo}'
            AND rec.año = '{año}'
        ORDER BY
            ops.ruta,
            ops.socio,
            ops.id,
            rec.nro_recibo
        """

    try:
        recibos = mant.run_query(instruccion, fetch="all")

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Cobrador.
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Socio/a.
              - Apellido y nombre.
              - Domicilio.
              - Ruta.
              - Importe.
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, title, 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # N° de cierre
            self.set_font('Arial', 'B', 10)
            self.cell(-77, 35, f'Cobrador: {nco}', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.cell(14, 5, 'Socio/a', 0, 0, 'L ')
            self.cell(65, 5, 'Apellido y Nombre', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(79, 5, 'Domicilio', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(15, 5, 'Ruta', 0, 0, 'L')
            pdf.cell(20, 5, 'Importe', 0, 1, 'L')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(3)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 3 cm from bottom
            self.set_y(-30)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 5, '* El asociado adeuda cuotas | En negrita: Cuotas a favor arregladas', 0, 1, 'L')
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 30)
    pdf.alias_nb_pages()
    pdf.add_page()

    for recibo in recibos:
        id_o, per, año, rut, c_f, nom_alt, dom_alt, nro, nom, dom, val_mant_bic, val_mant_nob = recibo

        nombre = nom_alt or nom if len(nom_alt or nom) < 36 else (nom_alt or nom)[:33] + '...'
        domicilio = dom_alt or dom if len(dom_alt or dom) < 36 else (dom_alt or dom)[:33] + '...'

        val_mant = val_mant_bic if facturacion == 'bicon' else val_mant_nob

        rec_impagos = rend.obtener_recibos_impagos_op(id_o)
        q_rec_impagos = len(rec_impagos)
        debe = 0

        if c_f < 0:
            debe += abs(c_f)
        debe += q_rec_impagos

        if q_rec_impagos:
            if rec_impagos[-1][2] == per:
                debe -= 1

        marca_debe = "*" if debe > 0 else ""

        pdf.set_font('Arial', '', 10)
        
        # Poniendo en negrita las operaciones con arreglo de cuotas a favor
        if id_o in ops_arregladas: pdf.set_font('Arial', 'B', 10)

        pdf.cell(14, 5, f'{nro}'.rjust(6, '0'), 0, 0, 'L ')
        pdf.cell(65,5, f'{nombre}{marca_debe}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(79, 5, f'{domicilio}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(10, 5, f'{rut}'.rjust(3, '0'), 0, 0, 'L')
        pdf.cell(20, 5, f'{val_mant:.2f}', 0, 1, 'R')
        imp_acu = imp_acu + float(val_mant)

    pdf.ln(2)
    pdf.cell(91, 5, '', 0, 0, 'L')
    pdf.cell(33, 5, 'Cantidad de recibos:', 'LTB', 0, 'L')
    pdf.cell(8, 5, f'{len(recibos)}', 'RTB', 0, 'R')
    pdf.cell(2, 5, '', 0, 0, 'L')
    pdf.cell(33, 5, 'Importe acumulado:', 'LTB', 0, 'L')
    pdf.cell(23, 5, f'$ {imp_acu:.2f}', 'RTB', 0, 'R')
    
    try:
        # Crea la carpeta si no existe
        if not os.path.isdir(mant.re_path(f'reports/recibos/{nco}')):
            os.mkdir(mant.re_path(f'reports/recibos/{nco}'))
    
        # Evita sobreescribir un archivo existente
        output_counter = 0
        prefix = "reimpresion_" if reimpresion else ""
        periodo = f"{per}_{año}".replace(' ', '')

        output_name = f"{prefix}listado_{periodo}.pdf"
    
        while os.path.isfile(mant.re_path(f'reports/recibos/{nco}/{output_name}')):
            output_counter += 1
            output_name = f'{prefix}listado_{periodo}_({output_counter}).pdf'
        
        pdf.output(mant.re_path(f'reports/recibos/{nco}/{output_name}'), 'F')

    ############ ABRIR REPORT ############

        
        ruta = mant.re_path(f'reports/recibos/{nco}/')
        arch = output_name.replace('(', '^(')
        os.chdir(ruta)
        os.system(arch)
        
        ruta = mant.MODULES_DIR
        os.chdir(ruta)

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return
          
############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################## RECIBOS DEB. AUT. ################################################
#################################################################################################################

def recibos_deb_aut(facturacion: str, periodo: str, año: str):
    """Genera un reporte en PDF para cada uno de los recibos correspondientes a operaciones
    con débito automático de una facturación y un período específicos y luego los guarda.

    :param facturacion: Tipo de facturación (bicon o nob).
    :type facturacion: str

    :param periodo: Período a imprimir.
    :type periodo: str

    :param año: Año del período a imprimir.
    :type año: str
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    counter = 0

    instruccion = f"""\
        SELECT
            rec.nro_recibo,
            ops.tarjeta,
            ops.nombre_alt,
            soc.nro_socio,
            soc.nombre,
            soc.cod_postal,
            pan.panteon,
            nic.piso,
            nic.fila,
            nic.numero,
            cat.categoria,
            cat.valor_mant_bicon,
            cat.valor_mant_nob
        FROM
            recibos rec
            JOIN operaciones ops ON rec.operacion = ops.id
            JOIN socios soc ON ops.socio = soc.nro_socio
            JOIN nichos nic ON ops.nicho = nic.codigo
            JOIN panteones pan ON nic.panteon = pan.id
            JOIN cat_nichos cat ON nic.categoria = cat.id
        WHERE
            ops.paga = 1
            AND soc.activo = 1
            AND rec.pago = 0
            AND ops.cobrador = 6
            AND ops.nicho IS NOT NULL
            AND ops.tarjeta IS NOT NULL
            AND rec.periodo = '{periodo}'
            AND rec.año = '{año}'
        ORDER BY
            ops.ruta,
            ops.socio,
            ops.id,
            rec.nro_recibo
        """

    try:
        recibos = mant.run_query(instruccion, fetch="all")

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############
    for recibo in recibos:
        pdf = FPDF()
    
        pdf.alias_nb_pages()
        pdf.add_page()

        ndr, tar, nom_alt, nro, nom, cod, pant, pis, fil, num, cat, val_mant_bic, val_mant_nob = recibo

        nombre = nom_alt or nom if len(nom_alt or nom) < 42 else (nom_alt or nom)[:39] + '...'

        val_mant = val_mant_bic if facturacion == 'bicon' else val_mant_nob

        _, _, _, t04 = rend.split_nro_tarjeta(tar)

        # Header NOB
        if facturacion == 'nob':
            # Logo
            pdf.image(mant.re_path('docs/logo_nob.jpg'), 11, 14, 10)
            pdf.set_font('Arial', 'I', 7)
            pdf.cell(190, 3.1, '', 0, 1, 'L')
            pdf.ln(1)
            
            # Línea 1
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(60, 3, "Club Atlético Newell's Old Boys", 0, 1, 'L')
            
            # Línea 2
            pdf.set_font('Arial', 'B', 6)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
            pdf.set_font('Arial', 'I', 6)
            pdf.cell(48, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
            pdf.cell(15, 3, '', 0, 1, 'L')
            
            # Línea 3
            pdf.set_font('Arial', 'B', 6)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(60, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 1, 'L')
            
            # Línea 4
            pdf.set_font('Arial', '', 6)
            pdf.cell(12, 3, '', 0, 0, 'L')
            pdf.cell(44, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
            pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
            pdf.ln(1)
        
        # Header Bicon
        else:
            # Línea 1
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(85, 5, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 1, 'C')
            
            # Línea 2
            pdf.cell(85, 5, 'Tel.: 430 9999 / 430 8800', 0, 1, 'C')
            
            # Línea 3
            pdf.cell(85, 5, 'CORDOBA 2915 - ROSARIO', 0, 1, 'C')
            
            # Línea 4
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(49, 5, '', 0, 0, 'L')
            pdf.cell(20, 5, 'Recibo nro.', 'LTB', 0, 'L')
            pdf.cell(16, 5, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
            pdf.ln(1)

        # Línea 1
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(13, 4, 'Socio/a: ', 'LTB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'TB', 0, 'L')
        pdf.cell(60, 4, f'{nombre}', 'TRB', 1, 'L')
        
        # Línea 2
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(85, 1, '', 0, 1, 'L')
        pdf.cell(19, 5, 'Cobrador/a: ', 'LT', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(5, 5, f'6'.rjust(2, '0'), 'T', 0, 'L')
        pdf.cell(41, 5, f'Débito Automático', 'T', 0, 'L')
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(10, 5, 'Tarjeta: ', 'T', 0, 'R')
        pdf.set_font('Arial', '', 9)
        pdf.cell(10, 5, f'{t04}', 'TR', 1, 'L')
        
        # Línea 3
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(17, 4, 'Categoría: ', 'LB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(68, 4, f'{cat}', 'BR', 1, 'L')
        pdf.ln(1)
        
        # Línea 4
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(19, 5, 'Cod. Nicho:', 'LTB', 0, 'L')
        pdf.set_font('Arial', '', 9)
        pdf.cell(66, 5, f'{cod}'.rjust(10, '0'), 'RTB', 1, 'L')
        pdf.ln(1)
        
        # Línea 5
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(44, 4, 'Panteón', 'LTR', 0, 'C')
        pdf.cell(13, 4, 'Piso', 'LTR', 0, 'C')
        pdf.cell(13, 4, 'Fila', 'LTR', 0, 'C')
        pdf.cell(15, 4, 'Nicho', 'LTR', 1, 'C')
        
        # Línea 6
        pdf.set_font('Arial', '', 9)
        pdf.cell(44, 5, f'{pant}', 'LBR', 0, 'C')
        pdf.cell(13, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
        pdf.cell(13, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
        pdf.cell(15, 5, f'{num}'.rjust(3, '0'), 'LBR', 1, 'C')
        pdf.ln(3)
        
        # Línea 7
        if periodo == "Enero - Febrero":
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(54, 5, f'{periodo} - {int(año)+1}', 'RTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(17, 5, f'$ {val_mant:.2f}', 'RTB', 1, 'C')
            pdf.ln(2)
        
        elif periodo == "Diciembre - Enero":
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(54, 5, f'{periodo} - {año}/{int(año[:2])+1}', 'RTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(17, 5, f'$ {val_mant:.2f}', 'RTB', 1, 'C')
            pdf.ln(2)
        
        else:
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(54, 5, f'{periodo} - {año}', 'RTB', 0, 'L')
            pdf.set_font('Arial', '', 9)
            pdf.cell(17, 5, f'$ {val_mant:.2f}', 'RTB', 1, 'C')
            pdf.ln(2)

        try:
            num_soc = f'{nro}'.rjust(6, '0')
            num_rec = f'{ndr}'.rjust(7, '0')
            
            # Crea la carpeta si no existe
            deb_aut_path = mant.re_path("reports/recibos/Débito automático")
            nombre_parseado = re.sub(r'[\\/:*?"<>|]', '_', nom)
            soc_path = f'{deb_aut_path}/{num_soc}-{nombre_parseado}'
            
            if not os.path.isdir(deb_aut_path):
                os.mkdir(deb_aut_path)

            if not os.path.isdir(f'{soc_path}'):
                os.mkdir(f'{soc_path}')
            
            pdf.output(f'{soc_path}/recibo-{num_rec}.pdf', 'F')

        except Exception as e:
            mant.manejar_excepcion_gral(e)
            print()
            return

        counter += 1
        mant.barra_progreso(counter, len(recibos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')

    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    print("\n")
    print("Abriendo listado...")
    print()

############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################## LISTADO DE RECIBOS DEB. AUT. ###########################################
#################################################################################################################
def listado_recibos_deb_aut(facturacion: str, periodo: str, año: str, ops_arregladas: list):
    """Genera un reporte en PDF que contiene un listado con la información de todos
    los recibos correspondientes a operaciones con débito automático de una facturación
    y un período específicos, luego lo guarda y lo abre con el programa predeterminado.
    
    Por último, genera, y guarda, un archivo TXT con la información y el formato
    correspondiente para realizar la carga de solicitudes de débito automático a 
    FiServ (presentación).

    El listado se reimprime con datos y valores actualizados.

    :param facturacion: Tipo de facturación (bicon o nob).
    :type facturacion: str

    :param periodo: Período a imprimir.
    :type periodo: str

    :param año: Año del período a imprimir.
    :type año: str

    :param ops_arregladas: Operaciones que se les necesitó arreglar sus cuotas a favor.
    :type ops_arregladas: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    dia = datetime.now().strftime('%d')
    mes = datetime.now().strftime('%m')
    imp_acu = float(0)
    nro_comercio_fiserv = str(mant.obtener_nro_comercio_fiserv())
    filler = " "
    contador_fiserv = 0
    val_total = 0

    instruccion = f"""\
        SELECT
            rec.operacion,
            ops.tarjeta,
            ops.cuotas_favor,
            ops.nombre_alt,
            soc.nro_socio,
            soc.nombre,
            cat.valor_mant_bicon,
            cat.valor_mant_nob
        FROM
            recibos rec
            JOIN operaciones ops ON rec.operacion = ops.id
            JOIN socios soc ON ops.socio = soc.nro_socio
            JOIN nichos nic ON ops.nicho = nic.codigo
            JOIN cat_nichos cat ON nic.categoria = cat.id
        WHERE
            ops.paga = 1
            AND soc.activo = 1
            AND rec.pago = 0
            AND ops.cobrador = 6
            AND ops.nicho IS NOT NULL
            AND ops.tarjeta IS NOT NULL
            AND rec.periodo = '{periodo}'
            AND rec.año = '{año}'
        ORDER BY
            ops.ruta,
            ops.socio,
            ops.id,
            rec.nro_recibo
        """
    
    try:
        recibos = mant.run_query(instruccion, fetch="all")

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############

    # Eliminar el archivo temporal de detalle de la presentación si es que existe
    pres_det_path = mant.re_path('reports/presentaciones_fiserv/temp/pres_det.txt')
    if os.path.isfile(pres_det_path):
        os.remove(pres_det_path)

    ############ FIN DE FUNCIONES ############

    ############ INICIO DE VARIABLES DEPENDIENTES ###########

    if periodo == 'Enero - Febrero':
        periodo_fiserv = f"1/{año[:2]} "
        vto_fiserv = f"2802{año[:2]}"
    
    elif periodo == 'Febrero - Marzo':
        periodo_fiserv = f"1/{año[:2]} "
        vto_fiserv = f"3103{año[:2]}"
    
    elif periodo == 'Marzo - Abril':
        periodo_fiserv = f"2/{año[:2]} "
        vto_fiserv = f"3004{año[:2]}"
    
    elif periodo == 'Abril - Mayo':
        periodo_fiserv = f"2/{año[:2]} "
        vto_fiserv = f"3105{año[:2]}"
    
    elif periodo == 'Mayo - Junio':
        periodo_fiserv = f"3/{año[:2]} "
        vto_fiserv = f"3006{año[:2]}"

    elif periodo == 'Junio - Julio':
        periodo_fiserv = f"3/{año[:2]} "
        vto_fiserv = f"3107{año[:2]}"

    elif periodo == 'Julio - Agosto':
        periodo_fiserv = f"4/{año[:2]} "
        vto_fiserv = f"3108{año[:2]}"

    elif periodo == 'Agosto - Septiembre':
        periodo_fiserv = f"4/{año[:2]} "
        vto_fiserv = f"3009{año[:2]}"

    elif periodo == 'Septiembre - Octubre':
        periodo_fiserv = f"5/{año[:2]} "
        vto_fiserv = f"3110{año[:2]}"
    
    elif periodo == 'Octubre - Noviembre':
        periodo_fiserv = f"5/{año[:2]} "
        vto_fiserv = f"3011{año[:2]}"
    
    elif periodo == 'Noviembre - Diciembre':
        periodo_fiserv = f"6/{año[:2]} "
        vto_fiserv = f"3112{año[:2]}"
    
    elif periodo == 'Diciembre - Enero':
        periodo_fiserv = f"6/{año[:2]} "
        vto_fiserv = f"3101{año[:2]+1}"

    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Socio/a.
              - Apellido y nombre.
              - Recibo.
              - Nro. Tarjeta.
              - Importe.
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'LISTADO DE RECIBOS EMITIDOS', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # N° de cierre
            self.set_font('Arial', 'B', 10)
            self.cell(-77, 35, f'Cobrador: Débito automático', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.cell(14, 5, 'Socio/a', 0, 0, 'L ')
            self.cell(98, 5, 'Apellido y Nombre', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(16, 5, ' Recibo', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(37, 5, '       Nro. Tarjeta', 0, 0, 'L')
            pdf.cell(22, 5, 'Importe', 0, 1, 'R')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(3)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 3 cm from bottom
            self.set_y(-30)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 5, '* El asociado adeuda cuotas', 0, 1, 'L')
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 30)
    pdf.alias_nb_pages()
    pdf.add_page()

    for recibo in recibos:
        # Variables individuales
        id_o, tar, c_f, nom_alt, nro, nom, val_mant_bic, val_mant_nob = recibo

        t01, t02, t03, t04 = rend.split_nro_tarjeta(tar)
        nombre = nom_alt or nom if len(nom_alt or nom) < 36 else (nom_alt or nom)[:33] + '...'
        val_mant = val_mant_bic if facturacion == 'bicon' else val_mant_nob
        ult_rec = str(rend.obtener_ult_rec_de_op(id_o)).rjust(7, '0')

        rec_impagos = rend.obtener_recibos_impagos_op(id_o)
        q_rec_impagos = len(rec_impagos)
        debe = 0

        if c_f < 0:
            debe += abs(c_f)
        debe += q_rec_impagos

        if q_rec_impagos:
            if rec_impagos[-1][2] == periodo:
                debe -= 1

        marca_debe = "*" if debe > 0 else ""

        pdf.set_font('Arial', '', 10)

        # Poniendo en negrita las operaciones con arreglo de cuotas a favor
        if id_o in ops_arregladas: pdf.set_font('Arial', 'B', 10)

        pdf.cell(14, 5, f'{nro}'.rjust(6, '0'), 0, 0, 'L ')
        pdf.cell(98,5, f'{nombre}{marca_debe}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(16, 5, f'{ult_rec}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(37, 5, f'{t01} {t02} {t03} {t04}', 0, 0, 'L')
        pdf.cell(22, 5, f'$ {val_mant:.2f}', 0, 1, 'R')
        imp_acu += float(val_mant)

    ########## Inscripción en el TXT de detalle de presentación para FiServ ##########

        # Crea las carpetas si no existen
        if os.path.isdir(mant.re_path('reports/presentaciones_fiserv')) == False:
            os.mkdir(mant.re_path('reports/presentaciones_fiserv'))
        
        if os.path.isdir(mant.re_path('reports/presentaciones_fiserv/temp')) == False:
            os.mkdir(mant.re_path('reports/presentaciones_fiserv/temp'))
        
        # Inscripción al detalle del TXT
        with open(mant.re_path('reports/presentaciones_fiserv/temp/pres_det.txt'), 'a') as detalle:
            detalle.write(f"\n{nro_comercio_fiserv}2{str(tar)}{str(id_o).rjust(12, '0')}00199902{str(int(val_mant)).rjust(9, '0')}00{periodo_fiserv} {vto_fiserv}{60*filler}")
        
        # Contador de cobros
        contador_fiserv += 1
        
        # Suma monto al total
        val_total += val_mant
    
                        ########## FIN DETALLE FISERV ##########

    pdf.ln(2)
    pdf.cell(91, 5, '', 0, 0, 'L')
    pdf.cell(33, 5, 'Cantidad de recibos:', 'LTB', 0, 'L')
    pdf.cell(8, 5, f'{len(recibos)}', 'RTB', 0, 'R')
    pdf.cell(2, 5, '', 0, 0, 'L')
    pdf.cell(33, 5, 'Importe acumulado:', 'LTB', 0, 'L')
    pdf.cell(23, 5, f'$ {imp_acu:.2f}', 'RTB', 0, 'R')
    
                
                ########## Generando TXT de presentación para FiServ ##########

    # Evita sobreescribir un archivo existente
    output_counter = 0
    output_name = f'presentacion_{año}-{mes}-{dia}.txt'

    while os.path.isfile(mant.re_path(f'reports/presentaciones_fiserv/{output_name}')):
        output_counter += 1
        output_name = f'presentacion_{año}-{mes}-{dia}_({output_counter}).txt'

    # Escritura de la cabecera de la presentacíon
    with open(mant.re_path(f'reports/presentaciones_fiserv/{output_name}'), 'w') as presentacion:
        presentacion.write(f"{nro_comercio_fiserv}1{dia}{mes}{año[:2]}{str(contador_fiserv).rjust(7, '0')}0{str(int(val_total)).rjust(12, '0')}00{91*filler}")

    # Escritura del detalle de la presentación
    try:
        with open(mant.re_path('reports/presentaciones_fiserv/temp/pres_det.txt'), 'r') as detalle:
            lista = detalle.readlines()

        for linea in lista:
            with open(mant.re_path(f'reports/presentaciones_fiserv/{output_name}'), 'a') as presentacion:
                presentacion.write(linea)

    except FileNotFoundError:
        print("         ERROR. No se encontraron recibos.")
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return
    
                        ########## FIN PRESENTACIÓN FISERV ##########
    
    try:
        # Crea la carpeta si no existe
        deb_aut_path = mant.re_path('reports/recibos/Débito automático')
        if not os.path.isdir(deb_aut_path):
            os.mkdir(deb_aut_path)

        # Evita sobreescribir un archivo existente
        output_counter = 0
        output_name = f'listado_recibos_{año}-{mes}-{dia}.pdf'

        while os.path.isfile(f'{deb_aut_path}/{output_name}'):
            output_counter += 1
            output_name = f'listado_recibos_{año}-{mes}-{dia}_({output_counter}).pdf'

        pdf.output(mant.re_path(f'{deb_aut_path}/{output_name}'), 'F')
        
    ############ ABRIR REPORT ############

        arch = f'listado_recibos_{año}-{mes}-{dia}.pdf'
        os.chdir(deb_aut_path)
        os.system(arch)

        ruta = mant.MODULES_DIR
        os.chdir(ruta)

    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################ RECIBOS  DE DOCUMENTOS #############################################
#################################################################################################################

def recibos_documentos() -> list:
    """Genera un reporte en PDF que contiene todos los recibos correspondientes
    a las operaciones recibidas como parámetro, luego lo guarda y lo abre con el
    programa predeterminado.

    Luego retorna una lista con todos los números de recibo generados.

    :rtype: list
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    cobradores = vent.obtener_cobradores()
    dia = datetime.now().strftime('%d')
    mes = datetime.now().strftime('%m')
    str_mes_sig, int_mes_sig = rend.obtener_mes_siguiente(mes)
    año = datetime.now().strftime('%Y')
    año2c = datetime.now().strftime('%y')
    lista_recibos = []
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############
    def days_between_abs(d1: date, d2: date) -> int:
        """Recibe dos fechas y retorna la diferencia absoluta en días entre ellas.

        :param d1: Fecha 1
        :type d1: datetime.date

        :param d2: Fecha 2
        :type d2: datetime.date

        :rtype: int
        """
        return abs(d2-d1).days

    def buscar_documentos() -> list:
        """Recupera de la base de datos toda la información de la tabla
        documentos y la retorna ordenada por ID de operación en una lista.
        
        :rtype: list
        """
        instruccion = f"SELECT * FROM documentos ORDER BY id_op"
        documentos = mant.run_query(instruccion, fetch="all")
        return documentos

    def obtener_añovar(mes: str, año: str) -> str:
        """Recibe el mes y el año y, si el mes recibido es Diciembre (12),
        retorna el año siguiente al recibido.

        :param mes: Mes (cadena, dos dígitos)
        :type mes: str

        :param año: Año (cadena, cuatro dígitos)
        :type año: str

        :rtype: str
        """
        añovar = int(año)
        if mes == '12':
            añovar += 1
        return str(añovar)

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############
    
    año_var = obtener_añovar(mes, año)

    ############ INICIO DE REPORT ############


    pdf = FPDF()

    pdf.set_margins(10, 0, 10)
    pdf.set_auto_page_break(True, 0)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    for i, cobrador in enumerate(cobradores):
        counter = 0
        print(f"{i+1}/{len(cobradores)}: ")
        id_cob, nom_cob = cobrador
        documentos = buscar_documentos()
        
        if len(documentos) == 0:
            mant.barra_progreso(counter, len(documentos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas - {i+1}/{len(cobradores)}')
        
        for documento in documentos:
            id_op, cant_cuotas, val_cuotas, ult_rec = documento
        
            if cant_cuotas > 0 and ult_rec != f'{int_mes_sig}-{año2c}':
                id_o, soc, nic, fac, cob, tar, rut, ult, u_a, fup, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = ctas.buscar_op_nro_operacion(id_op, True)
        
                if cob == id_cob:
                    # Variables individuales
                    try:
                        cod, pan, pis, fil, num, cat, ocu, fall = rend.obtener_datos_nicho(nic)
                    
                    except TypeError:
                        if 'Operaciones sin nicho' not in errores:
                            errores['Operaciones sin nicho'] = [str(id_o).rjust(7, '0')]
                    
                        else:
                            errores['Operaciones sin nicho'].append(str(id_o).rjust(7, '0'))
                    
                        continue
                    
                    nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = rend.obtener_datos_socio(soc)
                    id_c, cat, val_mant_bic, val_mant_nob = rend.obtener_categoria(cat)
                    pant = rend.obtener_panteon(pan)
                    nco = caja.obtener_nom_cobrador(cob)
                    fup_sep = str(fup).split("/")
                    fup_date = date(year = int(fup_sep[1]), month = int(fup_sep[0]), day = 1)
                    hoy = datetime.now().date()
                    cuenta = int(days_between_abs(hoy, fup_date)/365)
                    
                    if nom_alt != None:
                        nom = f"[{nom_alt}]"
                    
                    if dom_alt != None:
                        dom = f"[{dom_alt}]"
                    
                    parameters = str((id_o, f'Doc: {str_mes_sig}', año_var, 0))
                    query = f"INSERT INTO recibos (operacion, periodo, año, pago) VALUES {parameters}"
                    
                    mant.run_query(query)
                    
                    ndr = rend.obtener_nro_recibo()
                    lista_recibos.append(ndr)

                    # Header                                                   <----- HAY QUE PENSAR COMO PONER LA IMG DE ÑUL Y SEPARARLOS   
                    #                                                           (asegurarse que el header de bicon comentado coincida con éste)
                    # Línea 1
                    pdf.set_font('Arial', 'I', 7)
                    pdf.cell(190, 3.1, '', 0, 1, 'L')
                    pdf.cell(71, 3, '', 0, 0, 'L')
                    pdf.cell(25, 3, 'Talón para control', 0, 0, 'L')
                    pdf.cell(61, 3, '', 0, 0, 'L')
                    pdf.cell(31, 3, 'Talón para el contribuyente', 0, 1, 'L')
                    pdf.ln(1)
                    # Línea 2
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(7, 3, '', 0, 0, 'L')
                    pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 0, 'L')
                    pdf.cell(7, 3, '', 0, 0, 'L')
                    pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 1, 'L')
                    # Línea 3
                    pdf.cell(22, 3, '', 0, 0, 'L')
                    pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 0, 'L')
                    pdf.cell(18, 3, '', 0, 0, 'L')
                    pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 1, 'L')
                    # Línea 4
                    pdf.cell(19, 3, '', 0, 0, 'L')
                    pdf.cell(45, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
                    pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
                    pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
                    pdf.cell(20, 3, '', 0, 0, 'L')
                    pdf.cell(48, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
                    pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
                    pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
                    pdf.ln(1)

                    if fac == 'bicon': #                                                <----- HEADER DE BICON
                    #     # Línea 1
                    #     pdf.set_font('Arial', 'I', 7)
                    #     pdf.cell(190, 3.1, '', 0, 1, 'L')
                    #     pdf.cell(71, 3, '', 0, 0, 'L')
                    #     pdf.cell(25, 3, 'Talón para control', 0, 0, 'L')
                    #     pdf.cell(61, 3, '', 0, 0, 'L')
                    #     pdf.cell(31, 3, 'Talón para el contribuyente', 0, 1, 'L')
                    #     pdf.ln(1)
                    
                    #     # Línea 2
                    #     pdf.set_font('Arial', 'B', 8)
                    #     pdf.cell(7, 3, '', 0, 0, 'L')
                    #     pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 0, 'L')
                    #     pdf.cell(7, 3, '', 0, 0, 'L')
                    #     pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 1, 'L')
                    
                    #     # Línea 3
                    #     pdf.cell(22, 3, '', 0, 0, 'L')
                    #     pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 0, 'L')
                    #     pdf.cell(18, 3, '', 0, 0, 'L')
                    #     pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 1, 'L')
                    
                    #     # Línea 4
                    #     pdf.cell(19, 3, '', 0, 0, 'L')
                    #     pdf.cell(45, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
                    #     pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
                    #     pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
                    #     pdf.cell(20, 3, '', 0, 0, 'L')
                    #     pdf.cell(48, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
                    #     pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
                    #     pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
                    #     pdf.ln(1)
                        borrar = 'Esta variable es para poder contraer el if'
                    
                    elif fac == 'nob': #                                                <----- HEADER DE NOB
                    #     # Logo
                    #     pdf.image(mant.re_path('docs/logo_nob.jpg'), 11, 4, 10)
                    #     pdf.image(mant.re_path('docs/logo_nob.jpg'), 108, 4, 10)
                    #     pdf.set_font('Arial', 'I', 7)
                    #     pdf.cell(190, 3.1, '', 0, 1, 'L')
                    #     pdf.ln(1)
                    
                    #     # Línea 1
                    #     pdf.set_font('Arial', 'B', 8)
                    #     pdf.cell(12, 3, '', 0, 0, 'L')
                    #     pdf.cell(60, 3, "Club Atlético Newell's Old Boys", 0, 0, 'L')
                    #     pdf.set_font('Arial', 'I', 7)
                    #     pdf.cell(14, 3, 'Talón para control', 0, 0, 'L')
                    #     pdf.set_font('Arial', 'B', 8)
                    #     pdf.cell(24, 3, '', 0, 0, 'L')
                    #     pdf.cell(49, 3, "Club Atlético Newell's Old Boys", 0, 0, 'L')
                    #     pdf.set_font('Arial', 'I', 7)
                    #     pdf.cell(14, 3, 'Talón para el contribuyente', 0, 1, 'L')
                    
                    #     # Línea 2
                    #     pdf.set_font('Arial', 'B', 6)
                    #     pdf.cell(12, 3, '', 0, 0, 'L')
                    #     pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
                    #     pdf.set_font('Arial', 'I', 6)
                    #     pdf.cell(48, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
                    #     pdf.cell(15, 3, '', 0, 0, 'L')
                    #     pdf.set_font('Arial', 'B', 6)
                    #     pdf.cell(19, 3, '', 0, 0, 'L')
                    #     pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
                    #     pdf.set_font('Arial', 'I', 6)
                    #     pdf.cell(47, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
                    #     pdf.cell(15, 3, '', 0, 1, 'L')
                    
                    #     # Línea 3
                    #     pdf.set_font('Arial', 'B', 6)
                    #     pdf.cell(12, 3, '', 0, 0, 'L')
                    #     pdf.cell(60, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 0, 'L')
                    #     pdf.cell(38, 3, '', 0, 0, 'L')
                    #     pdf.cell(49, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 1, 'L')
                    
                    #     # Línea 4
                    #     pdf.set_font('Arial', '', 6)
                    #     pdf.cell(12, 3, '', 0, 0, 'L')
                    #     pdf.cell(52, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
                    #     pdf.set_font('Arial', 'B', 8)
                    #     pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
                    #     pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
                    #     pdf.set_font('Arial', '', 6)
                    #     pdf.cell(17, 3, '', 0, 0, 'L')
                    #     pdf.cell(51, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
                    #     pdf.set_font('Arial', 'B', 8)
                    #     pdf.set_font('Arial', 'B', 8)
                    #     pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
                    #     pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
                    #     pdf.ln(1)
                        borrar = 'Esta variable es para poder contraer el if'

                    # Línea 1
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(13, 4, 'Socio/a: ', 'LTB', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'TB', 0, 'L')
                    pdf.cell(68, 4, f'{nom}', 'TRB', 0, 'L')
                    pdf.cell(4, 4, '', 0, 0, 'L')
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(13, 4, 'Socio/a: ', 'LT', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'T', 0, 'L')
                    pdf.cell(68, 4, f'{nom}', 'TR', 1, 'L')
                    
                    # Línea 2
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(95, 1, '', 0, 0, 'L')
                    pdf.cell(2, 1, '', 0, 0, 'L')
                    pdf.cell(93, 1, '', 'LR', 1, 'L')
                    pdf.cell(19, 5, 'Cobrador/a: ', 'LT', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(5, 5, f'{cob}'.rjust(2, '0'), 'T', 0, 'L')
                    pdf.cell(69, 5, f'{nco}', 'RT', 0, 'L')
                    pdf.cell(4, 4, '', 0, 0, 'L')
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(17, 5, 'Domicilio: ', 'L', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(76, 5, f'{dom}', 'R', 1, 'L')
                    
                    # Línea 3
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(17, 4, 'Categoría: ', 'LB', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(46, 4, f'{cat}', 'B', 0, 'L')
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(9, 4, 'Ruta:', 'B', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(21, 4, f'{rut}'.rjust(3, '0'), 'RB', 0, 'L')
                    pdf.cell(4, 4, '', 0, 0, 'L')
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(17, 4, 'Localidad:', 'LB', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(76, 4, f'{loc} - {c_p}', 'BR', 1, 'L')
                    pdf.ln(1)
                    
                    # Línea 4
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(19, 5, 'Cod. Nicho:', 'LTB', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(44, 5, f'{cod}'.rjust(10, '0'), 'TB', 0, 'L')
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(7, 5, f'Op:', 'TB', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(23, 5, f'{id_o}'.rjust(7, "0"), 'RTB', 0, 'L')
                    pdf.cell(4, 5, '', 0, 0, 'L')
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(16, 5, 'Cobrador:', 'LTB', 0, 'L')
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(5, 5, f'{cob}'.rjust(2, '0'), 'TB', 0, 'L')
                    pdf.cell(72, 5, f'{nco}', 'RTB', 1, 'L')
                    pdf.ln(1)
                    
                    # Línea 5
                    pdf.set_font('Arial', 'B', 9)
                    pdf.cell(44, 4, 'Panteón', 'LTR', 0, 'C')
                    pdf.cell(16, 4, 'Piso', 'LTR', 0, 'C')
                    pdf.cell(16, 4, 'Fila', 'LTR', 0, 'C')
                    pdf.cell(17, 4, 'Nicho', 'LTR', 0, 'C')
                    pdf.cell(4, 4, '', 0, 0, 'C')
                    pdf.cell(29, 4, 'Categoría', 'LTR', 0, 'C')
                    pdf.cell(34, 4, 'Panteón', 'LTR', 0, 'C')
                    pdf.cell(9, 4, 'Piso', 'LTR', 0, 'C')
                    pdf.cell(9, 4, 'Fila', 'LTR', 0, 'C')
                    pdf.cell(12, 4, 'Nicho', 'LTR', 1, 'C')
                    
                    # Línea 6
                    pdf.set_font('Arial', '', 9)
                    pdf.cell(44, 5, f'{pant}', 'LBR', 0, 'C')
                    pdf.cell(16, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
                    pdf.cell(16, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
                    pdf.cell(17, 5, f'{num}'.rjust(3, '0'), 'LBR', 0, 'C')
                    pdf.cell(4, 5, '', 0, 0, 'C')
                    pdf.cell(29, 5, f'{cat}', 'LBR', 0, 'C')
                    pdf.cell(34, 5, f'{pant}', 'LBR', 0, 'C')
                    pdf.cell(9, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
                    pdf.cell(9, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
                    pdf.cell(12, 5, f'{num}'.rjust(3, '0'), 'LBR', 1, 'C')
                    pdf.ln(3)
                    
                    # Línea 7
                    if mes != "12":
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(12, 5, 'Cuota:', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(40, 5, f'{11-int(cant_cuotas)}/10  -  {str_mes_sig} de {año}', 'RTB', 0, 'L')
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(27, 5, f'$ {val_cuotas:.2f}', 'RTB', 0, 'R')
                        pdf.cell(4, 5, '', 0, 0, 'C')
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(12, 5, 'Cuota: ', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(40, 5, f'{11-int(cant_cuotas)}/10  -  {str_mes_sig} de {año}', 'RTB', 0, 'L')
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(27, 5, f'$ {val_cuotas:.2f}', 'RTB', 1, 'R')
                        pdf.ln(2)
                    
                    elif mes == "12":
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(12, 5, 'Cuota:', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(40, 5, f'{11-int(cant_cuotas)}/10  -  {str_mes_sig} de {int(año)+1}', 'RTB', 0, 'L')
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(27, 5, f'$ {val_cuotas:.2f}', 'RTB', 0, 'R')
                        pdf.cell(4, 5, '', 0, 0, 'C')
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(12, 5, 'Cuota: ', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(40, 5, f'{11-int(cant_cuotas)}/10  -  {str_mes_sig} de {int(año)+1}', 'RTB', 0, 'L')
                        pdf.set_font('Arial', 'B', 9)
                        pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
                        pdf.set_font('Arial', '', 9)
                        pdf.cell(27, 5, f'$ {val_cuotas:.2f}', 'RTB', 1, 'R')
                        pdf.ln(2)
                    
                    # Margen
                    pdf.cell(190, 17, ' ', 0, 1, 'L')

            counter += 1
            mant.barra_progreso(counter, len(documentos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas - {i+1}/{len(cobradores)}')
        print()
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')

    ############ GUARDAR REPORT ############
    try:
        # Crea la carpeta si no existe
        if not os.path.isdir(mant.re_path('reports/recibos/documentos')):
            os.mkdir(mant.re_path('reports/recibos/documentos'))

        # Evita la sobreescritura de archivos existentes
        output_counter = 0
        output_name = f'recibos_{año}-{mes}-{dia}.pdf'
        
        while os.path.isfile(mant.re_path(f'reports/recibos/documentos/{output_name}')):
            output_counter += 1
            output_name = f'recibos_{año}-{mes}-{dia}_({output_counter}).pdf'
        
        pdf.output(mant.re_path(f'reports/recibos/documentos/{output_name}'), 'F')

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión de los recibos se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo reporte. Cierre el archivo para continuar...")

        ruta = mant.re_path('reports/recibos/documentos')
        arch = output_name.replace('(', '^(')
        os.chdir(ruta)
        os.system(arch)

        ruta = mant.MODULES_DIR
        os.chdir(ruta)

    except UnboundLocalError:
        print()
        print("No se encontraron recibos impagos.")
        print()
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return
    
    return lista_recibos

############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################## LISTADO DE RECIBOS DE DOCUMENTOS #######################################
#################################################################################################################

def listado_recibos_documentos(lista_recibos: list):
    """Genera un reporte en PDF que contiene un listado con la información de todos
    los recibos recibidos como parámetro, luego lo guarda y lo abre con el programa
    predeterminado.

    :param lista_recibos: Lista de los recibos que se listarán.
    :type lista_recibos: list
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############

    dia = datetime.now().strftime('%d')
    mes = datetime.now().strftime('%m')
    str_mes_sig, int_mes_sig = rend.obtener_mes_siguiente(mes)
    año = datetime.now().strftime('%Y')
    año2c = datetime.now().strftime('%y')
    fecha = f"{mes}/{año}"
    hora = datetime.now().strftime('%H:%M')
    imp_acu = float(0)
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############
    def buscar_documento(id_op: int) -> tuple:
        """Recupera de la base de datos toda la información de un documento de
        una operación específica y la retorna en una tupla.

        :param id_op: ID de operación.
        :type id_op: int
        
        :rtype: tuple
        """
        instruccion = f"SELECT * FROM documentos WHERE id_op = {id_op}"
        documento = datos = mant.run_query(instruccion, fetch="one")
        return documento

    def obtener_añovar(mes: str, año: str) -> str:
        """Recibe el mes y el año y, si el mes recibido es Diciembre (12),
        retorna el año siguiente al recibido.

        :param mes: Mes (cadena, dos dígitos)
        :type mes: str

        :param año: Año (cadena, cuatro dígitos)
        :type año: str

        :rtype: str
        """
        añovar = int(año)
        if mes == '12':
            añovar += 1
        return str(añovar)

    def restar_cuota(id_op:int , cant_cuotas: int):
        """Modifica una operación específica registrándole una cuota menos que la
        cantidad recibida en el parámetro.

        :param id_op: ID de operación.
        :type id_op: int

        :param cant_cuotas: Cantidad de cuotas.
        :type cant_cuotas: int
        """
        instruccion = f"UPDATE documentos SET cant_cuotas = '{cant_cuotas-1}' WHERE id_op = '{id_op}'"
        mant.run_query(instruccion)

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############
    año_var = obtener_añovar(mes, año2c)

    ########### FIN DE VARIABLES DEPENDIENTES ############

    ############ INICIO DE FUNCIONES DEPENDIENTES ############
    def evitar_duplicado(id_op: int):
        """Actualiza el último recibo de una operación específica insertándole el
        mes actual. 

        Se utiliza para evitar emitir más de un recibo por período a la misma
        operación.

        :param id_op: ID de operación.
        :type id_op: int
        """
        ult_rec = f'{int_mes_sig}-{año_var}'
        instruccion = f"UPDATE documentos SET ult_rec = '{ult_rec}' WHERE id_op = '{id_op}'"
        mant.run_query(instruccion)

    ############ FIN DE FUNCIONES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Socio/a
              - Apellido y nombre
              - Cobrador
              - Ruta
              - Importe
            """            
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'LISTADO DE RECIBOS EMITIDOS', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # N° de cierre
            self.set_font('Arial', 'B', 10)
            self.cell(-77, 35, f'DE DOCUMENTOS', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.cell(14, 5, 'Socio/a', 0, 0, 'L ')
            self.cell(65, 5, 'Apellido y Nombre', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(75, 5, 'Domicilio', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(4, 5, 'Cob.', 0, 0, 'C')
            pdf.cell(1, 5, '', 0, 0, 'L')
            self.cell(15, 5, 'Ruta', 0, 0, 'L')
            pdf.cell(20, 5, 'Importe', 0, 1, 'L')
            self.cell(0, 1, '_________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(3)
            
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 3 cm from bottom
            self.set_y(-30)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 5, '* El asociado adeuda cuotas', 0, 1, 'L')
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()
    
    pdf.set_auto_page_break(True, 30)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    if len(lista_recibos) == 0:
        mant.barra_progreso(counter, len(lista_recibos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    for rec in lista_recibos:
        nro, ope, per, año, pag = rend.obtener_datos_recibo(rec)
        id_op, cant_cuotas, val_cuota, ult_rec = buscar_documento(ope)
        id_o, soc, nic, fac, cob, tar, rut, ult, u_a, fup, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = rend.obtener_datos_op(ope)
    
        try:
            cod, pan, pis, fil, num, cat, ocu, fall = rend.obtener_datos_nicho(nic)
    
        except TypeError:
            continue
    
        nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = rend.obtener_datos_socio(soc)
        
        if nom_alt != None:
            nom = f"[{nom_alt}]"
    
        if dom_alt != None:
            dom = f"[{dom_alt}]"
    
        pdf.set_font('Arial', '', 10)
        pdf.cell(14, 5, f'{nro}'.rjust(6, '0'), 0, 0, 'L ')
        pdf.cell(65,5, f'{nom}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(74, 5, f'{dom}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(4, 5, f'{cob}'.rjust(2, '0'), 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(10, 5, f'{rut}'.rjust(3, '0'), 0, 0, 'L')
        pdf.cell(20, 5, f'{val_cuota:.2f}', 0, 1, 'R')
        
        imp_acu += float(val_cuota)
    
        # Restando cuotas restantes
        restar_cuota(id_op, int(cant_cuotas))
    
        # Evitar duplicado de recibos
        evitar_duplicado(id_op)
        
        counter += 1
        mant.barra_progreso(counter, len(lista_recibos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    print()
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    pdf.ln(2)
    pdf.cell(91, 5, '', 0, 0, 'L')
    pdf.cell(33, 5, 'Cantidad de recibos:', 'LTB', 0, 'L')
    pdf.cell(8, 5, f'{len(lista_recibos)}', 'RTB', 0, 'R')
    pdf.cell(2, 5, '', 0, 0, 'L')
    pdf.cell(33, 5, 'Importe acumulado:', 'LTB', 0, 'L')
    pdf.cell(23, 5, f'$ {imp_acu:.2f}', 'RTB', 0, 'R')
    
    try:
        if not os.path.isdir(mant.re_path('reports/recibos/documentos')):
            os.mkdir(mant.re_path('reports/recibos/documentos'))
    
        output_counter = 0
        output_name = f'listado_recibos_{año}-{mes}-{dia}.pdf'
    
        while os.path.isfile(mant.re_path(f'reports/recibos/documentos/{output_name}')):
            output_counter += 1
            output_name = f'listado_recibos_{año}-{mes}-{dia}_({output_counter}).pdf'
    
        pdf.output(mant.re_path(f'reports/recibos/documentos/{output_name}'), 'F')

    ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión del listado se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo reporte. Cierre el archivo para continuar...")
    
        ruta = mant.re_path('reports/recibos/documentos/')
        arch = output_name.replace('(', '^(')
        os.chdir(ruta)
        os.system(arch)
    
        ruta = mant.MODULES_DIR
        os.chdir(ruta)
    
    except UnboundLocalError:
        print()
        print("No se encontraron recibos impagos.")
        print()
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return
          
############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################# REIMPRESIÓN DE RECIBO #############################################
#################################################################################################################

def reimpresion_recibo(ndr: int):
    """Genera un reporte en PDF que contiene un recibo en una única hoja a
    partir de su número de recibo, colocando en él, el valor actual del 
    mantenimiento. Luego lo guarda y lo abre con el programa predeterminado.

    :param ndr: Número de recibo.
    :type ndr: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    ndr, ope, per, año, pag = rend.obtener_datos_recibo(ndr)
    id_o, soc, nic, fac, cob, tar, rut, ult, u_a, fup, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = rend.obtener_datos_op(ope)
    
    try:
        cod, pan, pis, fil, num, cat, ocu, fall = rend.obtener_datos_nicho(nic)
    
    except TypeError:
        print()
        print("         ERROR. La operación no tiene nicho asociado.")
        print()
        return
    
    nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = rend.obtener_datos_socio(soc)
    id_c, cat, val_mant_bic, val_mant_nob = rend.obtener_categoria(cat)
    pant = rend.obtener_panteon(pan)
    nco = caja.obtener_nom_cobrador(cob)
    errores = {}
    
    if nom_alt != None:
        nom = f"[{nom_alt}]"
    
    if dom_alt != None:
        dom = f"[{dom_alt}]"
    
    if fac == 'bicon':
        val = val_mant_bic
    
    elif fac == 'nob':
        val = val_mant_nob
    
    if per[0:3] == 'Doc':
        val = rend.obtener_valor_doc(ope)
    
    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############
    pdf = FPDF()
    
    pdf.set_margins(10, 0, 10)
    pdf.set_auto_page_break(True, 0)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # Header
    if fac == 'bicon':
        # Línea 1   
        pdf.set_font('Arial', 'I', 7)
        pdf.cell(190, 3.1, '', 0, 1, 'L')
        pdf.cell(52, 3, '', 0, 0, 'L')
        pdf.cell(40, 3, 'Talón para control (REIMPRESIÓN)', 0, 0, 'L')
        pdf.cell(47, 3, '', 0, 0, 'L')
        pdf.cell(46, 3, 'Talón para el contribuyente (REIMPRESIÓN)', 0, 1, 'L')
        pdf.ln(1)
    
        # Línea 2
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(7, 3, '', 0, 0, 'L')
        pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 0, 'L')
        pdf.cell(7, 3, '', 0, 0, 'L')
        pdf.cell(87, 3, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 1, 'L')
    
        # Línea 3
        pdf.cell(22, 3, '', 0, 0, 'L')
        pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 0, 'L')
        pdf.cell(18, 3, '', 0, 0, 'L')
        pdf.cell(76, 3, 'Tel.: 430 9999 / 430 8800', 0, 1, 'L')
    
        # Línea 4
        pdf.cell(19, 3, '', 0, 0, 'L')
        pdf.cell(45, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
        pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
        pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
        pdf.cell(13, 3, '', 0, 0, 'L')
        pdf.cell(55, 3, 'CORDOBA 2915 - ROSARIO', 0, 0, 'L')
        pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
        pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
        pdf.ln(1)
    
    elif fac == 'nob':
        # Logo
        pdf.image(mant.re_path('docs/logo_nob.jpg'), 11, 4, 10)
        pdf.image(mant.re_path('docs/logo_nob.jpg'), 108, 4, 10)
        pdf.set_font('Arial', 'I', 7)
        pdf.cell(190, 3.1, '', 0, 1, 'L')
        pdf.ln(1)
    
        # Línea 1
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(60, 3, "Club Atlético Newell's Old Boys", 0, 0, 'L')
        pdf.set_font('Arial', 'I', 7)
        pdf.cell(14, 3, 'Talón para control', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(24, 3, '', 0, 0, 'L')
        pdf.cell(49, 3, "Club Atlético Newell's Old Boys", 0, 0, 'L')
        pdf.set_font('Arial', 'I', 7)
        pdf.cell(14, 3, 'Talón para el contribuyente', 0, 1, 'L')
    
        # Línea 2
        pdf.set_font('Arial', 'B', 6)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
        pdf.set_font('Arial', 'I', 6)
        pdf.cell(48, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
        pdf.cell(15, 3, '(REIMPRESIÓN)', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 6)
        pdf.cell(19, 3, '', 0, 0, 'L')
        pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
        pdf.set_font('Arial', 'I', 6)
        pdf.cell(47, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
        pdf.cell(15, 3, '(REIMPRESIÓN)', 0, 1, 'L')
    
        # Línea 3
        pdf.set_font('Arial', 'B', 6)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(60, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 0, 'L')
        pdf.cell(38, 3, '', 0, 0, 'L')
        pdf.cell(49, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 1, 'L')
    
        # Línea 4
        pdf.set_font('Arial', '', 6)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(52, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
        pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 0, 'R')
        pdf.set_font('Arial', '', 6)
        pdf.cell(17, 3, '', 0, 0, 'L')
        pdf.cell(51, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 8)
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
        pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
        pdf.ln(1)
    
    # Línea 1
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(13, 4, 'Socio/a: ', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'TB', 0, 'L')
    pdf.cell(68, 4, f'{nom}', 'TRB', 0, 'L')
    pdf.cell(4, 4, '', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(13, 4, 'Socio/a: ', 'LT', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'T', 0, 'L')
    pdf.cell(68, 4, f'{nom}', 'TR', 1, 'L')
    
    # Línea 2
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(95, 1, '', 0, 0, 'L')
    pdf.cell(2, 1, '', 0, 0, 'L')
    pdf.cell(93, 1, '', 'LR', 1, 'L')
    pdf.cell(19, 5, 'Cobrador/a: ', 'LT', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(5, 5, f'{cob}'.rjust(2, '0'), 'T', 0, 'L')
    pdf.cell(69, 5, f'{nco}', 'RT', 0, 'L')
    pdf.cell(4, 4, '', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(17, 5, 'Domicilio: ', 'L', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(76, 5, f'{dom}', 'R', 1, 'L')
    
    # Línea 3
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(17, 4, 'Categoría: ', 'LB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(46, 4, f'{cat}', 'B', 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(9, 4, 'Ruta:', 'B', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(21, 4, f'{rut}'.rjust(3, '0'), 'RB', 0, 'L')
    pdf.cell(4, 4, '', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(17, 4, 'Localidad:', 'LB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(76, 4, f'{loc} - {c_p}', 'BR', 1, 'L')
    pdf.ln(1)
    
    # Línea 4
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(19, 5, 'Cod. Nicho:', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(44, 5, f'{cod}'.rjust(10, '0'), 'TB', 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(7, 5, f'Op:', 'TB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(23, 5, f'{id_o}'.rjust(7, "0"), 'RTB', 0, 'L')
    pdf.cell(4, 5, '', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(16, 5, 'Cobrador:', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(5, 5, f'{cob}'.rjust(2, '0'), 'TB', 0, 'L')
    pdf.cell(72, 5, f'{nco}', 'RTB', 1, 'L')
    pdf.ln(1)
    
    # Línea 5
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(44, 4, 'Panteón', 'LTR', 0, 'C')
    pdf.cell(16, 4, 'Piso', 'LTR', 0, 'C')
    pdf.cell(16, 4, 'Fila', 'LTR', 0, 'C')
    pdf.cell(17, 4, 'Nicho', 'LTR', 0, 'C')
    pdf.cell(4, 4, '', 0, 0, 'C')
    pdf.cell(29, 4, 'Categoría', 'LTR', 0, 'C')
    pdf.cell(34, 4, 'Panteón', 'LTR', 0, 'C')
    pdf.cell(9, 4, 'Piso', 'LTR', 0, 'C')
    pdf.cell(9, 4, 'Fila', 'LTR', 0, 'C')
    pdf.cell(12, 4, 'Nicho', 'LTR', 1, 'C')
    
    # Línea 6
    pdf.set_font('Arial', '', 9)
    pdf.cell(44, 5, f'{pant}', 'LBR', 0, 'C')
    pdf.cell(16, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
    pdf.cell(16, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
    pdf.cell(17, 5, f'{num}'.rjust(3, '0'), 'LBR', 0, 'C')
    pdf.cell(4, 5, '', 0, 0, 'C')
    pdf.cell(29, 5, f'{cat}', 'LBR', 0, 'C')
    pdf.cell(34, 5, f'{pant}', 'LBR', 0, 'C')
    pdf.cell(9, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
    pdf.cell(9, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
    pdf.cell(12, 5, f'{num}'.rjust(3, '0'), 'LBR', 1, 'C')
    pdf.ln(3)
    
    # Línea 7
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    
    if per == 'Diciembre - Enero':
        pdf.cell(44, 5, f'{per} - {año}/{int(str(año)[-2:])+1}', 'RTB', 0, 'L')
    
    else:
        pdf.cell(44, 5, f'{per} - {año}', 'RTB', 0, 'L')
    
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(21, 5, f'$ {val:.2f}', 'RTB', 0, 'R')
    pdf.cell(4, 5, '', 0, 0, 'C')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(14, 5, 'Período: ', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    
    if per == 'Diciembre - Enero':
        pdf.cell(44, 5, f'{per} - {año}/{int(str(año)[-2:])+1}', 'RTB', 0, 'L')
    
    else:
        pdf.cell(44, 5, f'{per} - {año}', 'RTB', 0, 'L')
    
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(14, 5, 'Importe:', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(21, 5, f'$ {val:.2f}', 'RTB', 1, 'R')
    pdf.ln(2)
    
    # Línea 8
    pdf.cell(190, 17, ' ', 0, 1, 'L')
    
    # Margen
    pdf.cell(190, 13, ' ', 0, 1, 'L')
    
    try:
        if not os.path.isdir(mant.re_path('reports/temp')):
            os.mkdir(mant.re_path('reports/temp'))
        pdf.output(mant.re_path(f'reports/temp/recibo_{str(ndr).rjust(7, "0")}.pdf'), 'F')

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la reimpresión del recibo se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo recibo. Cierre el archivo para continuar...")
    
        ruta = mant.re_path('reports/temp')
        arch = f'recibo_{str(ndr).rjust(7, "0")}.pdf'
        os.chdir(ruta)
        os.system(arch)
    
        ruta = mant.MODULES_DIR
        os.chdir(ruta)
    
    except UnboundLocalError:
        print()
        print(f"No existe recibo con nro {str(ndr).rjust(7, '0')}.")
        print()
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return
        
############################################### FIN DE REPORT ###################################################




#################################################################################################################
################################################ RECIBO ADELANTOS ###############################################
#################################################################################################################

def recibo_adelanto(ndr: int, cobrador: int, periodo_h: str, año_h: str, valor_total: float | int):
    """Genera un reporte en PDF que contiene un único recibo de pago por adelanto
    de cuotas o pago de más de una cuota, luego lo guarda y lo abre con el programa
    predeterminado.

    :param ndr: Número de recibo.
    :type ndr: int

    :param cobrador: ID de cobrador.
    :type cobrador: int

    :param periodo_h: Período hasta el cual se abonó.
    :type periodo_h: str

    :param año_h: Año hasta el cual se abonó (cadena, 4 dígitos)
    :type año_h: str

    :param valor_total: Monto total abonado.
    :type valor_total: float or int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    ndr, ope, per, año, pag = rend.obtener_datos_recibo(ndr)
    id_o, soc, nic, fac, cob, tar, rut, ult, u_a, fup, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = rend.obtener_datos_op(ope)
    
    try:
        cod, pan, pis, fil, num, cat, ocu, fall = rend.obtener_datos_nicho(nic)
    
    except TypeError:
        print()
        print("         ERROR. La operación no tiene nicho asociado.")
        print()
        return
    
    nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = rend.obtener_datos_socio(soc)
    id_c, cat, val_mant_bic, val_mant_nob = rend.obtener_categoria(cat)
    panteon = rend.obtener_panteon(pan)
    nco = caja.obtener_nom_cobrador(cobrador)
    errores = {}
    
    if nom_alt != None:
        nom = f"[{nom_alt}]"
    
    if dom_alt != None:
        dom = f"[{dom_alt}]"    
    
    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############
    pdf = FPDF()
    
    pdf.alias_nb_pages()
    pdf.add_page()
    
    # Header
    # Bicon
    if fac == 'bicon':
        # Línea 1
        pdf.set_font('Arial', 'B', 11)
        pdf.cell(85, 5, 'ADMINISTRACIÓN de PANTEONES SOCIALES', 0, 1, 'C')

        # Línea 2
        pdf.cell(85, 5, 'Tel.: 430 9999 / 430 8800', 0, 1, 'C')

        # Línea 3
        pdf.cell(85, 5, 'CORDOBA 2915 - ROSARIO', 0, 1, 'C')

        # Línea 4
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(49, 5, '', 0, 0, 'L')
        pdf.cell(20, 5, 'Recibo nro.', 'LTB', 0, 'L')
        pdf.cell(16, 5, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
        pdf.ln(1)
    
    # NOB
    elif fac == 'nob':
        # Logo
        pdf.image(mant.re_path('docs/logo_nob.jpg'), 11, 14, 10)
        pdf.set_font('Arial', 'I', 7)
        pdf.cell(190, 3.1, '', 0, 1, 'L')
        pdf.ln(1)

        # Línea 1
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(60, 3, "Club Atlético Newell's Old Boys", 0, 1, 'L')
        
        # Línea 2
        pdf.set_font('Arial', 'B', 6)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(16, 3, 'Panteón Social', 0, 0, 'L')
        pdf.set_font('Arial', 'I', 6)
        pdf.cell(48, 3, '(Cementerio "El Salvador")', 0, 0, 'L')
        pdf.cell(15, 3, '', 0, 1, 'L')
        
        # Línea 3
        pdf.set_font('Arial', 'B', 6)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(60, 3, 'ADMINISTRACIÓN PANTEÓN SOCIAL', 0, 1, 'L')
        
        # Línea 4
        pdf.set_font('Arial', '', 6)
        pdf.cell(12, 3, '', 0, 0, 'L')
        pdf.cell(44, 3, 'Córdoba 2915 - Tel. 430 9999 / 8800', 0, 0, 'L')
        pdf.set_font('Arial', 'B', 8)
        pdf.cell(17, 3, 'Recibo nro.', 'LTB', 0, 'L')
        pdf.cell(12, 3, f'{ndr}'.rjust(7, '0'), 'RTB', 1, 'R')
        pdf.ln(1)

    # Línea 1
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(13, 4, 'Socio/a: ', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(12, 4, f'{nro}'.rjust(6, '0'), 'TB', 0, 'L')
    pdf.cell(60, 4, f'{nom}', 'TRB', 1, 'L')
    
    # Línea 2
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(85, 1, '', 0, 1, 'L')
    pdf.cell(16, 5, 'Domicilio: ', 'LT', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(69, 5, f'{dom}', 'TR', 1, 'L')
    
    # Línea 3
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(19, 5, 'Cobrador/a: ', 'L', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(5, 5, f'{cobrador}'.rjust(2, '0'), 0, 0, 'L')
    pdf.cell(39, 5, f'{nco}', 0, 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(10, 5, 'Ruta: ', 0, 0, 'R')
    pdf.set_font('Arial', '', 9)
    pdf.cell(12, 5, f'{rut}', 'R', 1, 'L')
    
    # Línea 4
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(17, 4, 'Categoría: ', 'LB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(68, 4, f'{cat}', 'BR', 1, 'L')
    pdf.ln(1)
    
    # Línea 5
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(19, 5, 'Cod. Nicho:', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    pdf.cell(43, 5, f'{cod}'.rjust(10, '0'), 'TB', 0, 'L')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(7, 5, f'Op:', 'TB', 0, 'TB')
    pdf.set_font('Arial', '', 9)
    pdf.cell(16, 5, f'{id_o}'.rjust(7, "0"), 'RTB', 1, 'L')
    pdf.ln(1)
    
    # Línea 6
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(44, 4, 'Panteón', 'LTR', 0, 'C')
    pdf.cell(13, 4, 'Piso', 'LTR', 0, 'C')
    pdf.cell(13, 4, 'Fila', 'LTR', 0, 'C')
    pdf.cell(15, 4, 'Nicho', 'LTR', 1, 'C')
    
    # Línea 7
    pdf.set_font('Arial', '', 9)
    pdf.cell(44, 5, f'{panteon}', 'LBR', 0, 'C')
    pdf.cell(13, 5, f'{pis}'.rjust(2, '0'), 'LBR', 0, 'C')
    pdf.cell(13, 5, f'{fil}'.rjust(2, '0'), 'LBR', 0, 'C')
    pdf.cell(15, 5, f'{num}'.rjust(3, '0'), 'LBR', 1, 'C')
    pdf.ln(3)
    
    # Línea 8
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(11, 5, 'Hasta: ', 'LTB', 0, 'L')
    pdf.set_font('Arial', '', 9)
    
    if periodo_h == 'Diciembre - Enero':
        pdf.cell(49, 5, f'{periodo_h} - {año_h}/{int(str(año_h)[-2:])+1}', 'RTB', 0, 'L')
    
    else:
        pdf.cell(49, 5, f'{periodo_h} - {año_h}', 'RTB', 0, 'L')
    
    pdf.set_font('Arial', '', 9)
    pdf.cell(25, 5, f'$ {valor_total:.2f}', 'RTB', 1, 'C')
    pdf.ln(2)
    
    # Guardado de archivo
    try:
        # Creación de la carpeta si no existe
        if not os.path.isdir(mant.re_path('reports/temp')):
            os.mkdir(mant.re_path('reports/temp'))
    
        pdf.output(mant.re_path(f'reports/temp/recibo_{str(ndr).rjust(7, "0")}.pdf'), 'F')

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión del recibo se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo recibo. Cierre el archivo para continuar...")
        
        ruta = mant.re_path('reports/temp/')
        arch = f'recibo_{str(ndr).rjust(7, "0")}.pdf'
        os.chdir(ruta)
        os.system(arch)
        
        ruta = mant.MODULES_DIR
        os.chdir(ruta)

    except UnboundLocalError:
        print()
        print(f"No existe recibo con nro {str(ndr).rjust(7, '0')}.")
        print()
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return
        
############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################### ESTADO DE CUENTA ################################################
#################################################################################################################

def report_estado_cta(nro_socio: int, nombre: str, dni: int, facturacion: str, domicilio: str, te_1: str, te_2: str, mail: str, c_p: int, localidad: str, act: int):
    """Genera un reporte en PDF que contiene el estado de cuenta de un asociado
    específico, luego lo guarda y lo abre con el programa predeterminado.

    :param nro_socio: ID de asociado.
    :type nro_socio: int

    :param nombre: Nombre del asosciado.
    :type nombre: str

    :param dni: Documento de identidad del asociado (sin puntos).
    :type dni: int

    :param facturacion: Tipo de facturación del asociado (bicon o nob).
    :type facturacion: str

    :param domicilio: Domicilio del asociado.
    :type domicilio: str

    :param te_1: Número de teléfono 1 del asociado.
    :type te_1: str

    :param te_2: Número de teléfono 2 del asociado.
    :type te_2: str

    :param mail: Dirección de correo electrónico del asociado.
    :type mail: str

    :param c_p: Código postal del asociado.
    :type c_p: int

    :param localidad: Localidad del asociado.
    :type localidad: str

    :param act: Estado del asociado (0 inactivo, 1 activo).
    :type act: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    errores = {}
    
    if act == 1:
        estado = 'ACTIVO'
    
    elif act == 0:
        estado = 'INACTIVO'
    
    if mail == None:
        mail = ''
    
    if te_1 == None:
        te_1 = ''
    
    if te_2 == None:
        te_2 = ''

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############
    
    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Datos del asociado (Arial Negrita 10p):
                - Número y nombre.
                - DNI.
                - E-Mail.
                - Teléfono 1.
                - Domicilio.
                - Teléfono 2.
                - Localidad.
                - Estado.
            """
            # Logo
            if facturacion == 'bicon':
                self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            elif facturacion == 'nob':
                self.image(mant.re_path('docs/logo_nob.jpg'), 14.5, 12, 13)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'ESTADO DE CUENTA', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.set_font('Arial', 'B', 10)
            self.cell(19, 5, 'Asociado:', 'LT', 0, 'L')
            self.cell(13, 5, f'{nro_socio}'.rjust(6, '0'), 'T', 0, 'L')
            self.cell(114, 5, f'-  {nombre}', 'T', 0, 'L')
            self.cell(15, 5, 'DNI:', 'T', 0, 'L')
            self.cell(29, 5, f'{dni}', 'TR', 1, 'R')
            self.cell(19, 5, 'E-Mail:', 'L', 0, 'L')
            self.cell(127, 5, f'{mail}', 0, 0, 'L')
            self.cell(20, 5, 'Tel 1:', 0, 0, 'L')
            self.cell(24, 5, f'{te_1}', 'R', 1, 'R')
            self.cell(19, 5, 'Domicilio:', 'L', 0, 'L')
            self.cell(127, 5, f'{domicilio}', 0, 0, 'L')
            self.cell(20, 5, 'Tel 2:', 0, 0, 'L')
            self.cell(24, 5, f'{te_2}', 'R', 1, 'R')
            self.cell(19, 5, 'Localidad:', 'LB', 0, 'L')
            self.cell(127, 5, f'{c_p} - {localidad}', 'B', 0, 'L')
            self.cell(14, 5, 'Estado:', 'B', 0, 'L')
            self.cell(30, 5, f'{estado}', 'BR', 1, 'R')
            # Line break
            self.ln(2)
        
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()
    
    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    operaciones = ctas.obtener_datos_op_por_nro_socio(nro_socio)
    
    for o in operaciones:
        id_op, soc, nic, fac, cob, tar, rut, ult, u_a, fec, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = o
        cob = caja.obtener_nom_cobrador(cob)
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(20, 5, 'Operación: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(16, 5, f'{id_op}'.rjust(7, '0'), 0, 0, 'L')
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(34, 5, 'Última cuota paga: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
    
        if ult == 'Diciembre - Enero':
            pdf.cell(50, 5, f'{ult} - {u_a}/{int(str(u_a)[-2:])+1}', 0, 0, 'L')
    
        else:
            pdf.cell(50, 5, f'{ult} - {u_a}', 0, 0, 'L')
    
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(19, 5, 'Cobrador: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(20, 5, f'{cob}', 0, 1, 'L')

        try:
            nic, pan, pis, fil, num, id_cat, ocu, fall = rend.obtener_datos_nicho(nic)
            id_cat, cat, val_mant_bic, val_mant_nob = rend.obtener_categoria(id_cat)

        except TypeError:
            nic = 0
            
        except Exception as e:
            mant.manejar_excepcion_gral(e)
            print()
            return

        if nic:
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(25, 5, 'Código Nicho: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(23, 5, f'{nic}', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(20, 5, 'Categoría: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(25, 5, f'{cat}', 0, 1, 'L')

    
        if op_cob != 0 or nom_alt != None or dom_alt != None:
            if op_cob == 0:
                op_cob = '-'
    
            if nom_alt == None:
                nom_alt = '-'
    
            if dom_alt == None:
                dom_alt = '-'
    
            pdf.ln(1)
            pdf.set_font('Arial', '', 8)
            pdf.cell(3, 3, '  (', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(16, 3, 'Op. Cobol: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 8)
            pdf.cell(8, 3, f'{op_cob}', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(2, 3, '', 0, 0, 'L')
            pdf.cell(27, 3, 'Nombre alternativo: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 8)
            pdf.cell(40, 3, f'{nom_alt}', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(2, 3, '', 0, 0, 'L')
            pdf.cell(29, 3, 'Domicilio alternativo: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 8)
            pdf.cell(40, 3, f'{dom_alt})', 0, 1, 'L')
            pdf.ln(1)
    
        if nic:
            pdf.ln(1)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(15, 5, '', 0, 0, 'L')
            pdf.cell(20, 5, 'Recibo', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(67, 5, 'Período', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(19, 5, 'Deuda', 0, 1, 'L')
        
            if c_f < 0:
                pdf.set_font('Arial', '', 10)
                pdf.cell(15, 5, '', 0, 0, 'L')
                pdf.cell(20, 5, f'N/D', 0, 0, 'L ')
                pdf.cell(1, 5, '', 0, 0, 'L')
        
                if fac == 'bicon':
                    pdf.cell(69, 5, f'Hasta Agosto-Septiembre 2022', 0, 0, 'L')
        
                if fac == 'nob':
                    pdf.cell(69, 5, f'Hasta Julio-Agosto 2022', 0, 0, 'L')
        
                pdf.cell(1, 5, '', 0, 0, 'L')
                pdf.cell(2, 5, '$', 0, 0, 'R')
                pdf.cell(23, 5, f'{float(ctas.deuda_vieja_por_op(id_op)):.2f}', 0, 1, 'R')
        
            recibos = ctas.buscar_recibos_por_op(id_op)
        
            for r in recibos:
                nro, ope, per, año, pag = r
                nic = ctas.buscar_nicho_por_op(ope)

                if fac == 'bicon':
                    val = val_mant_bic
        
                elif fac == 'nob':
                    val = val_mant_nob
        
                if per[0:3] == 'Doc':
                    val = rend.obtener_valor_doc(ope)
        
                pdf.set_font('Arial', '', 10)
                pdf.cell(15, 5, '', 0, 0, 'L')
                pdf.cell(20, 5, f'{nro}'.rjust(7, '0'), 0, 0, 'L ')
                pdf.cell(1, 5, '', 0, 0, 'L')
        
                if per == 'Diciembre - Enero':
                    pdf.cell(69, 5, f'{per} - {año}/{int(str(año)[-2:])+1}', 0, 0, 'L')
        
                else:
                    pdf.cell(69, 5, f'{per} - {año}', 0, 0, 'L')
        
                pdf.cell(1, 5, '', 0, 0, 'L')
                pdf.cell(2, 5, '$', 0, 0, 'R')
                pdf.cell(23, 5, f'{float(val):.2f}', 0, 1, 'R')
        else:
            pdf.ln(3)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(15, 5, '', 0, 0, 'L')
            pdf.cell(85, 5, 'LA OPERACIÓN NO TIENE UN NICHO ASIGNADO', 1, 1, 'L')
            pdf.ln(3)
    
        pdf.ln(3)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(40, 5, 'Deuda de operación: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(1, 5, '$', 0, 0, 'R')
        pdf.cell(0, 5, f'{ctas.deuda_por_op(id_op):.2f}', 0, 1, 'L')
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(27, 5, 'Cuotas a favor: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
    
        if c_f < 0:
            pdf.cell(0, 5, f'0', 0, 1, 'L')
    
        else:
            pdf.cell(0, 5, f'{c_f}', 0, 1, 'L')
    
        pdf.ln(2)
        pdf.cell(0, 2, '________________________________________', 0, 1, 'L')
        pdf.ln(2)
    
    pdf.set_font('Arial', 'B', 10)
    pdf.ln(10)
    pdf.cell(100, 5, '', 0, 0, 'L')
    pdf.cell(47, 5, 'Deuda total del asociado: ', 0, 0, 'L')
    pdf.set_font('Arial', '', 10)
    pdf.cell(1, 5, '$', 0, 0, 'R')
    pdf.cell(0, 5, f'{float(ctas.deuda_por_socio(nro_socio)):.2f}', 0, 1, 'L')
    
    # Export
    pdf.output(mant.re_path('reports/temp/estado_cta.pdf'), 'F')
        

    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')
            

    print("Abriendo reporte. Cierre el archivo para continuar...")
    print()
    
    ruta = mant.re_path('reports/temp')
    arch = f'estado_cta.pdf'
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)
           
############################################### FIN DE REPORT ###################################################




#################################################################################################################
####################################### ESTADO DE CUENTA P/EMAIL ################################################
#################################################################################################################

def report_estado_cta_mail(nro_socio: int, nombre: str, dni: int, facturacion: str, domicilio: str, te_1: str, te_2: str, mail: str, c_p: int, localidad: str, act: int):
    """Genera un reporte en PDF que contiene el estado de cuenta de un asociado
    específico para enviar por mail y luego lo guarda.

    :param nro_socio: ID de asociado.
    :type nro_socio: int

    :param nombre: Nombre del asosciado.
    :type nombre: str

    :param dni: Documento de identidad del asociado (sin puntos).
    :type dni: int

    :param facturacion: Tipo de facturación del asociado (bicon o nob).
    :type facturacion: str

    :param domicilio: Domicilio del asociado.
    :type domicilio: str

    :param te_1: Número de teléfono 1 del asociado.
    :type te_1: str

    :param te_2: Número de teléfono 2 del asociado.
    :type te_2: str

    :param mail: Dirección de correo electrónico del asociado.
    :type mail: str

    :param c_p: Código postal del asociado.
    :type c_p: int

    :param localidad: Localidad del asociado.
    :type localidad: str

    :param act: Estado del asociado (0 inactivo, 1 activo).
    :type act: int
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    errores = {}
    
    if act == 1:
        estado = 'ACTIVO'
    
    elif act == 0:
        estado = 'INACTIVO'

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############
    
    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############
    

    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Datos del asociado (Arial Negrita 10p):
                - Número y nombre.
                - DNI.
                - E-Mail.
                - Teléfono 1.
                - Domicilio.
                - Teléfono 2.
                - Localidad.
                - Estado.
            """
            # Logo
            if facturacion == 'bicon':
                self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            elif facturacion == 'nob':
                self.image(mant.re_path('docs/logo_nob.jpg'), 14.5, 12, 13)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'ESTADO DE CUENTA', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.set_font('Arial', 'B', 10)
            self.cell(19, 5, 'Asociado:', 'LT', 0, 'L')
            self.cell(13, 5, f'{nro_socio}'.rjust(6, '0'), 'T', 0, 'L')
            self.cell(114, 5, f'-  {nombre}', 'T', 0, 'L')
            self.cell(15, 5, 'DNI:', 'T', 0, 'L')
            self.cell(29, 5, f'{dni}', 'TR', 1, 'R')
            self.cell(19, 5, 'E-Mail:', 'L', 0, 'L')
            self.cell(127, 5, f'{mail}', 0, 0, 'L')
            self.cell(20, 5, 'Tel 1:', 0, 0, 'L')
            self.cell(24, 5, f'{te_1}', 'R', 1, 'R')
            self.cell(19, 5, 'Domicilio:', 'L', 0, 'L')
            self.cell(127, 5, f'{domicilio}', 0, 0, 'L')
            self.cell(20, 5, 'Tel 2:', 0, 0, 'L')
            if te_2 == None or te_2 == "":
                self.cell(24, 5, f'', 'R', 1, 'R')
            else:
                self.cell(24, 5, f'{te_2}', 'R', 1, 'R')
            self.cell(19, 5, 'Localidad:', 'LB', 0, 'L')
            self.cell(127, 5, f'{c_p} - {localidad}', 'B', 0, 'L')
            self.cell(14, 5, 'Estado:', 'B', 0, 'L')
            self.cell(30, 5, f'{estado}', 'BR', 1, 'R')
            # Line break
            self.ln(2)
        
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    operaciones = ctas.obtener_datos_op_por_nro_socio(nro_socio)
    
    for o in operaciones:
        id_op, soc, nic, fac, cob, tar, rut, ult, u_a, fec, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = o
        cob = caja.obtener_nom_cobrador(cob)
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(20, 5, 'Operación: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(16, 5, f'{id_op}'.rjust(7, '0'), 0, 0, 'L')
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(34, 5, 'Última cuota paga: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
    
        if ult == 'Diciembre - Enero':
            pdf.cell(50, 5, f'{ult} - {u_a}/{int(str(u_a)[-2:])+1}', 0, 0, 'L')
    
        else:
            pdf.cell(50, 5, f'{ult} - {u_a}', 0, 0, 'L')
    
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(19, 5, 'Cobrador: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(20, 5, f'{cob}', 0, 1, 'L')

        try:
            nic, pan, pis, fil, num, id_cat, ocu, fall = rend.obtener_datos_nicho(nic)
            id_cat, cat, val_mant_bic, val_mant_nob = rend.obtener_categoria(id_cat)

        except TypeError:
            nic = 0
            
        except Exception as e:
            mant.manejar_excepcion_gral(e)
            print()
            return

        if nic:
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(25, 5, 'Código Nicho: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(23, 5, f'{nic}', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(20, 5, 'Categoría: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(25, 5, f'{cat}', 0, 1, 'L')

    
        if op_cob != 0 or nom_alt != None or dom_alt != None:
            if op_cob == 0:
                op_cob = '-'
    
            if nom_alt == None:
                nom_alt = '-'
    
            if dom_alt == None:
                dom_alt = '-'
    
            pdf.ln(1)
            pdf.set_font('Arial', '', 8)
            pdf.cell(3, 3, '  (', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(16, 3, 'Op. Cobol: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 8)
            pdf.cell(8, 3, f'{op_cob}', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(2, 3, '', 0, 0, 'L')
            pdf.cell(27, 3, 'Nombre alternativo: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 8)
            pdf.cell(40, 3, f'{nom_alt}', 0, 0, 'L')
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(2, 3, '', 0, 0, 'L')
            pdf.cell(29, 3, 'Domicilio alternativo: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 8)
            pdf.cell(40, 3, f'{dom_alt})', 0, 1, 'L')
            pdf.ln(1)
    
        if nic:
            pdf.ln(1)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(15, 5, '', 0, 0, 'L')
            pdf.cell(20, 5, 'Recibo', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(67, 5, 'Período', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(19, 5, 'Deuda', 0, 1, 'L')
        
            if c_f < 0:
                pdf.set_font('Arial', '', 10)
                pdf.cell(15, 5, '', 0, 0, 'L')
                pdf.cell(20, 5, f'N/D', 0, 0, 'L ')
                pdf.cell(1, 5, '', 0, 0, 'L')
        
                if fac == 'bicon':
                    pdf.cell(69, 5, f'Hasta Agosto-Septiembre 2022', 0, 0, 'L')
        
                if fac == 'nob':
                    pdf.cell(69, 5, f'Hasta Julio-Agosto 2022', 0, 0, 'L')
        
                pdf.cell(1, 5, '', 0, 0, 'L')
                pdf.cell(2, 5, '$', 0, 0, 'R')
                pdf.cell(23, 5, f'{float(ctas.deuda_vieja_por_op(id_op)):.2f}', 0, 1, 'R')
        
            recibos = ctas.buscar_recibos_por_op(id_op)
        
            for r in recibos:
                nro, ope, per, año, pag = r
                nic = ctas.buscar_nicho_por_op(ope)

                if fac == 'bicon':
                    val = val_mant_bic
        
                elif fac == 'nob':
                    val = val_mant_nob
        
                if per[0:3] == 'Doc':
                    val = rend.obtener_valor_doc(ope)
        
                pdf.set_font('Arial', '', 10)
                pdf.cell(15, 5, '', 0, 0, 'L')
                pdf.cell(20, 5, f'{nro}'.rjust(7, '0'), 0, 0, 'L ')
                pdf.cell(1, 5, '', 0, 0, 'L')
        
                if per == 'Diciembre - Enero':
                    pdf.cell(69, 5, f'{per} - {año}/{int(str(año)[-2:])+1}', 0, 0, 'L')
        
                else:
                    pdf.cell(69, 5, f'{per} - {año}', 0, 0, 'L')
        
                pdf.cell(1, 5, '', 0, 0, 'L')
                pdf.cell(2, 5, '$', 0, 0, 'R')
                pdf.cell(23, 5, f'{float(val):.2f}', 0, 1, 'R')
        else:
            pdf.ln(3)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(15, 5, '', 0, 0, 'L')
            pdf.cell(85, 5, 'LA OPERACIÓN NO TIENE UN NICHO ASIGNADO', 1, 1, 'L')
            pdf.ln(3)
    
        pdf.ln(3)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(40, 5, 'Deuda de operación: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(1, 5, '$', 0, 0, 'R')
        pdf.cell(0, 5, f'{ctas.deuda_por_op(id_op):.2f}', 0, 1, 'L')
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(27, 5, 'Cuotas a favor: ', 0, 0, 'L')
        pdf.set_font('Arial', '', 10)
    
        if c_f < 0:
            pdf.cell(0, 5, f'0', 0, 1, 'L')
    
        else:
            pdf.cell(0, 5, f'{c_f}', 0, 1, 'L')
    
        pdf.ln(2)
        pdf.cell(0, 2, '________________________________________', 0, 1, 'L')
        pdf.ln(2)
    
    pdf.set_font('Arial', 'B', 10)
    pdf.ln(10)
    pdf.cell(100, 5, '', 0, 0, 'L')
    pdf.cell(47, 5, 'Deuda total del asociado: ', 0, 0, 'L')
    pdf.set_font('Arial', '', 10)
    pdf.cell(1, 5, '$', 0, 0, 'R')
    pdf.cell(0, 5, f'{float(ctas.deuda_por_socio(nro_socio)):.2f}', 0, 1, 'L')
    
    # Export

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    pdf.output(mant.re_path('reports/temp/estado_cta_mail.pdf'), 'F')
        
############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################### LISTADO DE MOROSOS DETALLADO ##########################################
#################################################################################################################

def report_morosos_det():
    """Genera un reporte en PDF que contiene un listado con todos los estados de
    cuentas de aquellos socios que tengan operaciones en estado moroso. Luego lo
    guarda y lo abre con el programa predeterminado.
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    morosos = ctas.obtener_op_morosos()
    deuda_total_morosos = 0
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############
    def lista_morosos_limpia(morosos: list) -> set:
        """Recibe una lista de todos las operaciones morosas y retorna
        un set con sus respectivos asociados.

        :param morosos: Lista de operaciones morosas.
        :type morosos: list

        :rtype: set
        """
        lista_morosos = []

        for i in morosos:
            i_d, soc, nic, fac, cob, tar, rut, ult, u_a, fec_u_p, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = i
            lista_morosos.append(soc)
        
        return set(lista_morosos)

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############
    lista_limpia_morosos = lista_morosos_limpia(morosos)

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'LISTADO DETALLADO DE MOROSOS', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # Line break
            self.ln(22)
      
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 2.5 cm from bottom
            self.set_y(-25)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()
    
    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    for i in lista_limpia_morosos:
        nro_soc, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = rend.obtener_datos_socio(i)
    
        if act == 1:
            estado = 'ACTIVO'
    
        elif act == 0:
            estado = 'INACTIVO'
    
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(19, 5, 'Asociado:', 'LT', 0, 'L')
        pdf.cell(13, 5, f'{nro_soc}'.rjust(6, '0'), 'T', 0, 'L')
        pdf.cell(114, 5, f'-  {nom}', 'T', 0, 'L')
        pdf.cell(20, 5, 'Estado:', 'T', 0, 'L')
        pdf.cell(24, 5, f'{estado}', 'TR', 1, 'R')
        pdf.cell(19, 5, 'Domicilio:', 'L', 0, 'L')
        pdf.cell(127, 5, f'{dom}', 0, 0, 'L')
        pdf.cell(15, 5, 'Tel 1:', 0, 0, 'L')
        pdf.cell(29, 5, f'{te_1}', 'R', 1, 'R')
        pdf.cell(19, 5, 'Localidad:', 'LB', 0, 'L')
        pdf.cell(127, 5, f'{c_p} - {loc}', 'B', 0, 'L')
        pdf.cell(15, 5, 'Tel 2:', 'B', 0, 'L')
        
        if te_2 == None or te_2 == "":
            pdf.cell(29, 5, f'- ', 'BR', 1, 'R')
        
        else:
            pdf.cell(29, 5, f'{te_2}', 'BR', 1, 'R')        
        
        pdf.ln(2)
        
        for x in ctas.obtener_datos_op_por_nro_socio(nro_soc):
            id_op, soc, nic, fac, cob, tar, rut, ult, u_a, fec_u_p, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = x
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(10, 5, '', 0, 0, 'L')
            pdf.cell(20, 5, 'Operación: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(16, 5, f'{id_op}'.rjust(7, '0'), 0, 0, 'L')
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(34, 5, 'Última cuota paga: ', 0, 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(40, 5, f'{ult} - {u_a}', 0, 1, 'L')
            
            if op_cob != 0 or nom_alt != None or dom_alt != None:
                if op_cob == 0:
                    op_cob = '-'
            
                if nom_alt == None:
                    nom_alt = '-'
            
                if dom_alt == None:
                    dom_alt = '-'
            
                pdf.ln(1)
                pdf.set_font('Arial', '', 8)
                pdf.cell(3, 3, '  (', 0, 0, 'L')
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(16, 3, 'Op. Cobol: ', 0, 0, 'L')
                pdf.set_font('Arial', '', 8)
                pdf.cell(8, 3, f'{op_cob}', 0, 0, 'L')
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(2, 3, '', 0, 0, 'L')
                pdf.cell(27, 3, 'Nombre alternativo: ', 0, 0, 'L')
                pdf.set_font('Arial', '', 8)
                pdf.cell(40, 3, f'{nom_alt}', 0, 0, 'L')
                pdf.set_font('Arial', 'B', 8)
                pdf.cell(2, 3, '', 0, 0, 'L')
                pdf.cell(29, 3, 'Domicilio alternativo: ', 0, 0, 'L')
                pdf.set_font('Arial', '', 8)
                pdf.cell(40, 3, f'{dom_alt})', 0, 1, 'L')
                pdf.ln(1)
            
            pdf.ln(1)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(20, 5, '', 0, 0, 'L')
            pdf.cell(15, 5, 'Recibo', 0, 0, 'L ')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(11, 5, 'Cód. nicho', 0, 0, 'L')
            pdf.cell(12, 5, '', 0, 0, 'L')
            pdf.cell(38, 5, 'Período', 0, 0, 'L')
            pdf.cell(1, 5, '', 0, 0, 'L')
            pdf.cell(23, 5, 'Deuda', 0, 1, 'L')
            
            recibos = ctas.buscar_recibos_por_op(id_op)
            
            for r in recibos:
                nro_rec, ope, per, año, pag = r
                nic = ctas.buscar_nicho_por_op(ope)
            
                try:
                    cod, pan, pis, fil, num, cat, ocu, fall = rend.obtener_datos_nicho(nic)
                    id_cat, cat, val_mant_bic, val_mant_nob = rend.obtener_categoria(cat)
            
                except TypeError:
                    if 'Operaciones sin nicho' not in errores:
                        errores['Operaciones sin nicho'] = [str(id_op).rjust(7, '0')]
            
                    else:
                        errores['Operaciones sin nicho'].append(str(id_op).rjust(7, '0'))
            
                    continue
            
                if fac == 'bicon':
                    val = val_mant_bic
            
                elif fac == 'nob':
                    val = val_mant_nob
            
                if per[0:3] == 'Doc':
                    val = rend.obtener_valor_doc(ope)
            
                pdf.set_font('Arial', '', 10)
                pdf.cell(20, 5, '', 0, 0, 'L')
                pdf.cell(15, 5, f'{nro_rec}'.rjust(6, '0'), 0, 0, 'L ')
                pdf.cell(1, 5, '', 0, 0, 'L')
                pdf.cell(11, 5, f'{ctas.buscar_nicho_por_op(ope)}'.rjust(10, '0'), 0, 0, 'L')
                pdf.cell(12, 5, '', 0, 0, 'L')
                pdf.cell(40, 5, f'{per} - {año}', 0, 0, 'L')
                pdf.cell(1, 5, '', 0, 0, 'L')
                pdf.cell(2, 5, '$', 0, 0, 'R')
                pdf.cell(21, 5, f'{float(val):.2f}', 0, 1, 'R')
            
            pdf.set_font('Arial', 'B', 10)
            pdf.ln(3)
            pdf.cell(10, 5, '', 0, 0, 'L')
            pdf.cell(40, 5, 'Deuda de operación: ', 'LTB', 0, 'L')
            pdf.set_font('Arial', '', 10)
            pdf.cell(1, 5, '$', 'TB', 0, 'R')
            pdf.cell(23, 5, f'{float(ctas.deuda_por_op(id_op)):.2f}', 'RTB', 1, 'L')
            pdf.ln(2)

        pdf.set_font('Arial', 'B', 10)
        pdf.ln(10)
        pdf.cell(100, 5, '', 0, 0, 'L')
        pdf.cell(57, 5, 'Deuda total del asociado: ', 'LTB', 0, 'L')
        pdf.set_font('Arial', '', 10)
        pdf.cell(1, 5, '$', 'TB', 0, 'R')
        pdf.cell(23, 5, f'{float(ctas.deuda_por_socio(nro_soc)):.2f}', 'RTB', 1, 'L')

        deuda_total_morosos = deuda_total_morosos + ctas.deuda_por_socio(nro_soc)

        # Line break
        pdf.ln(3)

        counter += 1
        mant.barra_progreso(counter, len(lista_limpia_morosos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 5, f'DEUDA TOTAL MOROSOS: $ {deuda_total_morosos:.2f}', 1, 1, 'R')
    
    # Export
    pdf.output(mant.re_path('reports/temp/morosos.pdf'), 'F')
        

    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Cierre el archivo para continuar...")
    print()

    ruta = mant.re_path('reports/temp')
    arch = f'morosos.pdf'
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)
    
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
           
############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################### LISTADO DE MOROSOS COMPRIMIDO ##########################################
#################################################################################################################

def report_morosos_comp():
    """Genera un reporte en PDF que contiene un listado con la información de todos los
    socios que tengan operaciones en estado moroso, luego lo guarda y lo abre con el
    programa predeterminado.
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    morosos = ctas.obtener_op_morosos()
    deuda_total_morosos = 0
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############

        
    ############ INICIO DE FUNCIONES ############
    def lista_morosos_limpia(morosos: list) -> set:
        """Recibe una lista de todos las operaciones morosas y retorna
        un set con sus respectivos asociados.

        :param morosos: Lista de operaciones morosas.
        :type morosos: list

        :rtype: set
        """
        lista_morosos = []

        for i in morosos:
            i_d, soc, nic, fac, cob, tar, rut, ult, u_a, fec_u_p, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = i
            lista_morosos.append(soc)
        
        return set(lista_morosos)


    def reducir_ult_pag(ult: str, u_a: str) -> str:
        """Recibe un período bimestral y un año y retorna el mismo periodo
        reducido a siete caracteres y el año.

        :param ult: Período del último pago.
        :type ult: str

        :param u_a: Año del último pago.
        :type u_a: str

        :rtype: str
        """
        if ult == 'Enero - Febrero':
            ult_pag = f'Ene-Feb/{u_a}'
        
        elif ult == 'Febrero - Marzo':
            ult_pag = f'Feb-Mar/{u_a}'
        
        elif ult == 'Marzo - Abril':
            ult_pag = f'Mar-Abr/{u_a}'
        
        elif ult == 'Abril - Mayo':
            ult_pag = f'Abr-May/{u_a}'
        
        elif ult == 'Mayo - Junio':
            ult_pag = f'May-Jun/{u_a}'
        
        elif ult == 'Junio - Julio':
            ult_pag = f'Jun-Jul/{u_a}'
        
        elif ult == 'Julio - Agosto':
            ult_pag = f'Jul-Ago/{u_a}'
        
        elif ult == 'Agosto - Septiembre':
            ult_pag = f'Ago-Sep/{u_a}'
        
        elif ult == 'Septiembre - Octubre':
            ult_pag = f'Sep-Oct/{u_a}'
        
        elif ult == 'Octubre - Noviembre':
            ult_pag = f'Oct-Nov/{u_a}'
        
        elif ult == 'Noviembre - Diciembre':
            ult_pag = f'Nov-Dic/{u_a}'
        
        elif ult == 'Diciembre - Enero':
            ult_pag = f'Dic-Ene/{u_a}-{int(u_a[2:4])+1}'
            
        else:
            ult_pag = 'ERROR'
        
        return ult_pag

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############
    # lista_limpia_morosos = lista_morosos_limpia(morosos)

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - Número de operación
              - Número de socio
              - Apellido y nombre
              - Teléfonos
              - Último pago
              - Deuda
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            
            # Title
            self.set_font('Arial', 'B', 15)
            self.cell(0, 20, 'LISTADO COMPRIMIDO DE MOROSOS', 1, 0, 'C')
            
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')

            # Line break
            self.ln(22)

            # Table's header
            self.set_font('Arial', 'B', 9)
            self.cell(16, 5, 'N° OPER.', 0, 0, 'L ')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(13, 5, 'N° SOC.', 0, 0, 'L')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(53, 5, 'APELLIDO Y NOMBRE (*)', 0, 0, 'L')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(51, 5, 'TELÉFONOS', 0, 0, 'L')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(31, 5, 'ÚLTIMO PAGO', 0, 0, 'L')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(21, 5, 'DEUDA', 0, 1, 'L')
            self.cell(0, 1, '___________________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
        
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 3 cm from bottom
            self.set_y(-30)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 5, '* El asociado se encuentra INACTIVO', 0, 1, 'L')
            self.cell(0, 1, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()

    for i in morosos:
        id_op, soc, nic, fac, cob, tar, rut, ult, u_a, fec_u_p, mor, c_f, u_r, paga, op_cob, nom_alt, dom_alt = i
        nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = rend.obtener_datos_socio(soc)

        if nom_alt != None:
            nom = f"[{nom_alt}]"

        pdf.set_font('Arial', '', 9)
        pdf.cell(16, 5, f'{id_op}'.rjust(7, '0'), 0, 0, 'L ')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(13, 5, f'{nro}'.rjust(6, '0'), 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')

        if act == 1:
            pdf.cell(53, 5, f'{nom}', 0, 0, 'L')

        elif act == 0:
            pdf.cell(53, 5, f'{nom} *', 0, 0, 'L')

        pdf.cell(1, 5, '', 0, 0, 'L')

        if te_2 == None or te_2 == "":
            pdf.cell(51, 5, f'{te_1}', 0, 0, 'L')

        else:
            pdf.cell(51, 5, f'{te_1} / {te_2}', 0, 0, 'L')

        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(31, 5, f'{reducir_ult_pag(ult, u_a)}', 0, 0, 'L')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(21, 5, f'$ {float(ctas.deuda_por_op(id_op)):.2f}', 0, 1, 'R')

        deuda_total_morosos = deuda_total_morosos + ctas.deuda_por_op(id_op)

        counter += 1
        mant.barra_progreso(counter, len(morosos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')

    pdf.ln(3)
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 5, f'DEUDA TOTAL MOROSOS: $ {deuda_total_morosos:.2f}', 1, 1, 'R')
    
    # Export
    pdf.output(mant.re_path('reports/temp/morosos-comp.pdf'), 'F')
        

    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Cierre el archivo para continuar...")
    print()

    ruta = mant.re_path('reports/temp')
    arch = f'morosos-comp.pdf'
    os.chdir(ruta)
    os.system(arch)

    ruta = mant.MODULES_DIR
    os.chdir(ruta)

    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
           
############################################### FIN DE REPORT ###################################################




#################################################################################################################
########################################## LISTADO DE SOCIOS EN EXCEL ###########################################
#################################################################################################################

def report_excel_socios():
    """Genera un reporte en XLSX que contiene todos los asociados registrados
    en la base de datos y su respectiva inormación, luego lo guarda y lo abre
    con el programa predeterminado.
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = datetime.today().strftime("%Y-%m-%d")
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############

    instruccion = "SELECT * FROM socios"
    datos = mant.run_query(instruccion, fetch="all")

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    print("Generando archivo Excel")
    print()
    print("Progreso: ")

    wb = Workbook()
    ws = wb.active
    ws.title = "Socios"
    ws.append(('Nro. de socio', 'Nombre', 'DNI', 'Teléfono 1', 'Teléfono 2', 'Mail', 'Domicilio', 'Localidad', 'Código postal', 'Fecha de nacimiento', 'Fecha de alta', 'Activo'))
    
    for x in datos:
        nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act = x
        nro = f'{nro}'.rjust(6, '0')
    
        if act == 1:
            act = "SI"
    
        elif act == 2:
            act = "NO"
    
        datos_socio = (nro, nom, dni, te_1, te_2, mail, dom, loc, c_p, f_n, f_a, act)
        ws.append(datos_socio)
    
        counter += 1
        mant.barra_progreso(counter, len(datos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')

    print()
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    # Export
    try:
        wb.save(mant.re_path(f"reports/excel/listado_socios-{fecha}.xlsx"))
        print()
        print("Archivo creado exitosamente.")

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo reporte. Cierre el archivo para continuar...")
        print()

        ruta = mant.re_path('reports/excel')
        arch = f'listado_socios-{fecha}.xlsx'
        os.chdir(ruta)
        os.system(arch)
        
        ruta = mant.MODULES_DIR
        os.chdir(ruta)
    
    except PermissionError:
        print()
        print("         ERROR. El archivo no pudo ser creado porque se encuentra en uso. Ciérrelo y vuelva a intentarlo.")
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################ MODIFICACIONES DE CAJA #############################################
#################################################################################################################

def report_excel_modif_caja():
    """Genera un reporte en XLSX que contiene todos los registros de caja que
    hayan sido modificados, lo guarda y lo abre con el programa predetermiando.
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha = datetime.today().strftime("%Y-%m-%d")
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############

    instruccion = "SELECT * FROM historial_caja"
    datos = mant.run_query(instruccion, fetch="all")

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    print("Generando archivo Excel")
    print()
    print("Progreso: ")

    wb = Workbook()
    ws = wb.active
    ws.title = "Socios"
    ws.append(('ID', 'Categoría', 'Descripción', 'Transacción', 'Ingreso', 'Egreso', 'Observación', 'Usuario que modificó', 'Fecha y hora de modificación'))
    
    for x in datos:
        id_caj, cat, des, tra, ing, egr, obs, usr, fyh = x
        id_u, nom, ape, tel, dom, use, pas, pri, act =  mant.buscar_usuario_por_id(usr)
        datos_reg = (id_caj, cat, des, tra, ing, egr, obs, use, fyh)
        ws.append(datos_reg)
    
        counter += 1
    
        mant.barra_progreso(counter, len(datos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')    
    print()

    # Export
    try:
        wb.save(mant.re_path(f"reports/excel/modificaciones_de_caja-{fecha}.xlsx"))
        print()
        print("Archivo creado exitosamente.")

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo reporte. Cierre el archivo para continuar...")
        print()

        ruta = mant.re_path('reports/excel')
        arch = f'modificaciones_de_caja-{fecha}.xlsx'
        os.chdir(ruta)
        os.system(arch)
        
        ruta = mant.MODULES_DIR
        os.chdir(ruta)

    except PermissionError:
        print()
        print("         ERROR. El archivo no pudo ser creado porque se encuentra en uso. Ciérrelo y vuelva a intentarlo.")
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

############################################### FIN DE REPORT ###################################################




#################################################################################################################
######################################## MOVIMIENTOS DEB. AUT. EN EXCEL #########################################
#################################################################################################################

def report_deb_aut(mes: str, año: str):
    """Genera un reporte en XLSX que contiene todos los registros de pagos
    con débito autmático realizados en un mes específico, luego lo guarda
    y lo abre con el programa predeterminado.

    :param mes: Mes (cadena, dos dígitos).
    :type mes: str

    :param año: Año (cadena, cuatro dígitos).
    :type año: str
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############

    instruccion = f"SELECT * FROM debitos_automaticos WHERE mes = '{mes}' AND año = '{año}'"
    datos = mant.run_query(instruccion, fetch="all")

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    print("Generando archivo Excel")
    print()
    print("Progreso: ")
    

    wb = Workbook()
    ws = wb.active
    ws.title = "Socios"
    ws.append(('ID', 'Categoría', 'Socio', 'Operación', 'Ingreso', 'Observación', 'Fecha', 'Usuario'))
    
    for x in datos:
        id_debaut, cat, socio, oper, ingr, obs, dia, mes, año, user = x
        fecha = f'{dia}/{mes}/{año}'
        usuario = mant.buscar_usuario_por_id(user)[5]
        datos_debaut = (id_debaut, cat, socio, oper, ingr, obs, fecha, usuario)
        ws.append(datos_debaut)
    
        counter += 1
        mant.barra_progreso(counter, len(datos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    print()

    # Export
    try:
        wb.save(mant.re_path(f"reports/excel/listado_socios-{fecha_hoy}.xlsx"))
        print()
        print("Archivo creado exitosamente.")

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo reporte. Cierre el archivo para continuar...")
        print()
    
        ruta = mant.re_path('reports/excel')
        arch = f'listado_socios-{fecha_hoy}.xlsx'
        os.chdir(ruta)
        os.system(arch)
    
        ruta = mant.MODULES_DIR
        os.chdir(ruta)
    
    except PermissionError:
        print()
        print("         ERROR. El archivo no pudo ser creado porque se encuentra en uso. Ciérrelo y vuelva a intentarlo.")
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################# LISTADO DE COBRADORES #############################################
#################################################################################################################

def report_cobradores():
    """Genera un reporte en PDF que contiene un listado con todos los cobradores
    ordenados por su ID, luego lo guarda y lo abre con el programa predeterminado.
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############
    
    instruccion = f"SELECT * FROM cobradores ORDER BY id"
    cobradores = mant.run_query(instruccion, fetch="all")

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - ID del cobrador
              - Nombre del cobrador
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'LISTADO DE COBRADORES', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.set_font('Arial', 'B', 9)
            self.cell(10, 5, 'ID', 0, 0, 'L ')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(90, 5, 'COBRADOR', 0, 1, 'L')
            self.cell(0, 1, '___________________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
        
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 3 cm from bottom
            self.set_y(-30)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 7, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()

    for i in cobradores:
        i_d, cob = i
        pdf.cell(10, 5, f'{i_d}', 0, 0, 'L ')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(90, 5, f'{cob}', 0, 1, 'L')
    
        counter += 1
        mant.barra_progreso(counter, len(cobradores), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    print()
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    # Export
    pdf.output(mant.re_path('reports/temp/listado_cobradores.pdf'), 'F')
    
    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Cierre el archivo para continuar...")
    print()
    
    ruta = mant.re_path('reports/temp')
    arch = f'listado_cobradores.pdf'
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)

############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################# LISTADO DE COBRADORES #############################################
#################################################################################################################

def report_panteones():
    """Genera un reporte en PDF que contiene un listado con todos los panteones
    ordenados por su ID, luego lo guarda y lo abre con el programa predeterminado.
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    fecha = caja.obtener_fecha()
    hora = datetime.now().strftime('%H:%M')
    counter = 0
    errores = {}

    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############

    instruccion = f"SELECT * FROM panteones ORDER BY id"
    panteones = mant.run_query(instruccion, fetch="all")

    ############ FIN DE FUNCIONES ############


    ############ INICIO DE VARIABLES DEPENDIENTES ############

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    class PDF(FPDF):
        # Page header
        def header(self):
            """Escribe un encabezado para cada página del documento.

            Contenido:
            - Logo de la empresa.
            - Título del documento (Arial Negrita 15p).
            - Fecha y hora (Arial 10p).
            - Nombres de columnas (Arial Negrita 15p):
              - ID del panteón.
              - Nombre del panteón.
            """
            # Logo
            self.image(mant.re_path('docs/logo_bicon.jpg'), 14.5, 12, 15)
            # Arial bold 15
            self.set_font('Arial', 'B', 15)
            # Title
            self.cell(0, 20, 'LISTADO DE PANTEONES', 1, 0, 'C')
            # Arial 10
            self.set_font('Arial', '', 10)
            # Fecha
            self.cell(0, 35, f'{fecha} - {hora} hs', 0, 0, 'R')
            # Line break
            self.ln(22)
            self.set_font('Arial', 'B', 9)
            self.cell(10, 5, 'ID', 0, 0, 'L ')
            self.cell(1, 5, '', 0, 0, 'L')
            self.cell(90, 5, 'PANTEÓN', 0, 1, 'L')
            self.cell(0, 1, '___________________________________________________________________________________________________________', 0, 1, 'C')
            self.ln(5)
        
        # Page footer
        def footer(self):
            """Escribe un pie para cada página del documento.

            Contenido:
            - Número y total de páginas.
            - Nombre y versión de Morella.
            - Logo de MF! Soluciones Informáticas.
            """
            # Position at 3 cm from bottom
            self.set_y(-30)
            # Arial italic 8
            self.set_font('Arial', 'I', 8)
            self.cell(0, 7, '_______________________________________________________________________________________________', 0, 1, 'C')
            # Page number
            self.cell(0, 10, 'Página ' + str(self.page_no()) + ' de {nb}', 0, 0, 'C')
            # Firma
            self.set_font('Arial', 'I', 8)
            self.cell(-10, 10, f'Reporte generado en *MORELLA v{mant.SHORT_VERSION}* by ', 0, 0, 'R')
            self.image(mant.re_path('docs/mf_logo.jpg'), 190, 274, 8)

    # Instantiation of inherited class
    pdf = PDF()

    pdf.set_auto_page_break(True, 25)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    for i in panteones:
        i_d, cob = i
        pdf.cell(10, 5, f'{i_d}', 0, 0, 'L ')
        pdf.cell(1, 5, '', 0, 0, 'L')
        pdf.cell(90, 5, f'{cob}', 0, 1, 'L')
    
        counter += 1
        mant.barra_progreso(counter, len(panteones), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    print()
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    # Export
    pdf.output(mant.re_path('reports/temp/listado_panteones.pdf'), 'F')
        

    ############ ABRIR REPORT ############

    if errores:
        print('\n\n\n\n')
        print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
        print()
        pprint(errores)
        print('\n\n\n\n')

    print("Abriendo reporte. Cierre el archivo para continuar...")
    print()
    
    ruta = mant.re_path('reports/temp')
    arch = f'listado_panteones.pdf'
    os.chdir(ruta)
    os.system(arch)
    
    ruta = mant.MODULES_DIR
    os.chdir(ruta)

############################################### FIN DE REPORT ###################################################




#################################################################################################################
############################################# LISTADO ÚLTIMO RECIBO #############################################
#################################################################################################################

def report_ult_recibo(cobrador: int, facturacion: str):
    """Genera un reporte en Excel que contiene los datos de los últimos recibos
    impagos de cada operación de un cobrador y una facturación específicos.

    :param cobrador: id del cobrador
    :type cobrador: int

    :param facturacion: Categoría de facturación (bicon / nob)
    :type facturacion: str
    """
    ############ INICIO DE VARIABLES INDEPENDIENTES ############
    counter = 0
    errores = {}
    n_cob = caja.obtener_nom_cobrador(cobrador)
    title = f"Últimos recibos de {n_cob}"
    if len(title) > 31:
        title = f'Ult.rec.{n_cob}'[:31]
    
    ############ FIN DE VARIABLES INDEPENDIENTES ############


    ############ INICIO DE FUNCIONES ############
    def ult_rec_imp(cobrador: int, facturacion: str) -> list:
        """Retorna los siguientes datos de los últimos recibos impagos de todas
        las operaciones de un cobrador y una facturación específicos:

        - Número de socio
        - Nombre del asociado o nombre alternativo si posee
        - Número de recibo
        - Período y año
        - Importe

        :param cobrador: id del cobrador
        :type cobrador: int

        :param facturacion: Categoría de facturación (bicon / nob)
        :type facturacion: str
        
        :rtype: list

        """
        instruccion = f"""
        SELECT o.socio, s.nombre, o.nombre_alt, r.nro_recibo, r.periodo, r.año, cn.valor_mant_bicon, cn.valor_mant_nob
        FROM recibos r
        JOIN operaciones o
        ON r.operacion = o.id
        JOIN socios s
        ON o.socio = s.nro_socio
        JOIN nichos n
        ON o.nicho = n.codigo
        JOIN cat_nichos cn
        ON n.categoria = cn.id
        WHERE o.cobrador = {cobrador}
        AND r.pago = 0
        AND o.facturacion = '{facturacion}'
        AND nro_recibo IN(
            SELECT MAX(nro_recibo) 
            FROM recibos
            GROUP BY operacion
        )
        ORDER BY r.nro_recibo;
        """
        ultimos_recibos = mant.run_query(instruccion, fetch="all")
        return ultimos_recibos

    ############ FIN DE FUNCIONES ############
        

    ############ INICIO DE VARIABLES DEPENDIENTES ############
    ultimos_recibos = ult_rec_imp(cobrador, facturacion)

    ########### FIN DE VARIABLES DEPENDIENTES ############


    ############ INICIO DE REPORT ############
    print("Generando archivo Excel")
    print()

    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(('Número de socio', 'Nombre asociado','Número de recibo', 'Período', 'Importe'))
    
    for recibo in ultimos_recibos:
        nro_soc, nombre, nombre_alt, nro_rec, periodo, año, val_mant_bic, val_mant_nob = recibo
    
        if nombre_alt:
            nombre = nombre_alt
    
        if periodo == 'Diciembre - Enero':
            periodo_recibo = f"{periodo} {año}/{int(str(año)[-2:])+1}"
    
        else:
            periodo_recibo = f"{periodo} {año}"
    
        if facturacion == 'bicon':
            importe = val_mant_bic
    
        elif facturacion == 'nob':
            importe = val_mant_nob
    
        datos_recibo = (nro_soc, nombre, nro_rec, periodo_recibo, importe)
        ws.append(datos_recibo)
    
        counter += 1
        mant.barra_progreso(counter, len(ultimos_recibos), titulo=f'Morella v{mant.VERSION} - MF! Soluciones informáticas')
    
    os.system(f'TITLE Morella v{mant.VERSION} - MF! Soluciones informáticas')
    print()

    # Export
    try:
        if ' ' in n_cob:
            n_cob = n_cob.replace(' ', '_')
    
        wb.save(mant.re_path(f"reports/excel/ultimos_recibos_{n_cob}.xlsx"))
        print()
        print("Archivo creado exitosamente.")

        ############ ABRIR REPORT ############

        if errores:
            print('\n\n\n\n')
            print('     ATENCIÓN! Durante la emisión del reporte se produjeron los siguientes errores:')
            print()
            pprint(errores)
            print('\n\n\n\n')

        print("Abriendo reporte. Cierre el archivo para continuar...")
        print()
    
        ruta = mant.re_path('reports/excel')
        arch = f'ultimos_recibos_{n_cob}.xlsx'
        os.chdir(ruta)
        os.system(arch)
    
        ruta = mant.MODULES_DIR
        os.chdir(ruta)
    
    except PermissionError:
        print()
        print("         ERROR. El archivo no pudo ser creado porque se encuentra en uso. Ciérrelo y vuelva a intentarlo.")
    except Exception as e:
        mant.manejar_excepcion_gral(e)
        print()
        return

############################################### FIN DE REPORT ###################################################
